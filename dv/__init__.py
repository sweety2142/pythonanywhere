from email import message
import os
import re
import ssl
import sys
import operator
import copy
from flask import Flask,request, session, redirect, url_for, abort,render_template, flash, make_response, jsonify
from jinja2 import TemplateNotFound
from dv.database import *
from datetime import datetime, timedelta
from pprint import pprint as pp

import sys, os, time
from bs4 import BeautifulSoup
#from selenium import webdriver as wd

#from selenium.webdriver.support import expected_conditions as EC
#from selenium.webdriver.common.by import By
import openpyxl 
#from selenium.webdriver.support.ui import Select
#from selenium.webdriver.common.keys import Keys
import requests
#import win32com.client

# check list 
# 1. order 위치 이동 (o) 
# 2. log email 삭제 (o)  20200831 loginpage.html
# 3. pdf 폰트 (o)
# 4. modify  (o) - 내가 봤을땐 이상이 없다 
# 5. order commision 안되는 부분 확인하기 (o)  -  내가 봤을 땐 이상이 없다 
# 6. com total 문제 (o)
# 7. order 번호 제일 큰 번호로 

from flask_compress import Compress
compress = Compress()

def create_app():
    app = Flask(__name__)
    #Compress(app)
    compress.init_app(app)
    app.config['COMPRESS_LEVEL'] = 9
    #Cfg
    from dv.config import Cfg
    app.config.from_object(Cfg)
    app.secret_key = app.config['SECRET_KEY']
    
    #log
    from dv.logger import Log
    log_filepath = os.path.join(app.config['LOGGER_FILE'])
    Log.init(log_filepath=log_filepath)
    
    #blueprint
    from dv.dev_blueprint import admin_bp, users_bp, dashboards_bp
    app.register_blueprint(admin_bp)
    app.register_blueprint(users_bp, url_prefix='/admin')
    app.register_blueprint(dashboards_bp, url_prefix='/admin')
        
    USER_LIST = app.config['USER_PATH']
    LAYOUT_URL= app.config['TEMPLATE']['layout']
    LOG_FILE  = app.config['LOG_FILE']
    MAIN_URL  = app.config['TEMPLATE']['main']
    MAIN_URL_NAME = MAIN_URL.split('.')[0]
    SIGNUP_URL = app.config['TEMPLATE']['signup']
    LOGIN_URL  = app.config['TEMPLATE']['login']
    LOG_ACTIVITY=app.config['LOG_ACTIVITY']
    DB_CONT     = app.config['DB']
    TEMPLATE    = app.config['TEMPLATE']
    CHECK_COMPANY = 'DB'

    #session['company'] = CHECK_COMPANY
    
    #function
    def update_session(login_info):
        session['logged_in'] = True
        session['num']=login_info[0]
        session['name']=login_info[1]
        session['mail']=login_info[2]
        session['pw']=login_info[3]
        session['admin']=login_info[6]
        session['company']=CHECK_COMPANY
        
#        if login_info[1] in app.config['ADMIN_USER']:
#          print("Session ADMIN mode")
#          session['admin'] = True
    
    def session_logout(destination, flash_msg):
        resp = make_response(redirect(url_for(destination)))
        resp.set_cookie('cookie_id','',max_age=0) #remove cookie
        num = session['num']        
        session.pop('logged_in',None)
        session.pop('name',None)
        session.pop('pw',None)
        session.pop('mail',None)
        session.pop('cookie_id',None)
        session.pop('admin',None)
        session.pop('company',None)
        update_db(USER_LIST, update_query('user_list',['cookie_id'],'num'),[None,num])
        session.pop('num',None)
        # insert new 
        session.clear()
        return resp
        
    def get_form_data(keys):
        form_data = dict()
        for key in keys:
            val = request.form.getlist(key)
            form_data[key] = val
        return form_data
            
    def check_user_info():
        qdb = query_db(USER_LIST, select_query('user_list'))
        for row in qdb:
            if session['name'] == row[1]:
                #print(session['pw'])
                #print(session['mail'])
                #print(row[3])
                #print(row[2])
                if session['pw'] == row[3] and session['name'] == row[1] and session['mail'] == row[2]:
                    return True
        return False
        
    def logins():
        try:
            error = None
            qdb=query_db(USER_LIST,select_query('user_list'))

#            session['company'] = CHECK_COMPANY


            if request.method=='POST':
                #umail = request.form['mail']
                umail = ''
                name = request.form['name']
                pw= request.form['password']

                
                
                #pp(qdb)
                for row in qdb:
                    if name == row[1] and umail == row[2] and pw == row[3]:
                        record_logging(LOG_FILE, request.remote_addr, name+' '+umail, 'IN','logging')
                        update_session(row)
                        resp = make_response(redirect(url_for(MAIN_URL_NAME)))
                        if request.form.get('remember'):
                            cookie_id=set_random_id(20)
                            resp.set_cookie('cookie_id', cookie_id, max_age=app.config['REMEMBER_PERIOD'])
                            update_db(USER_LIST, update_query('user_list',['cookie_id'],'num'),[cookie_id,row[0]])
                            flash('You were logged in')
                            return make_response(redirect(url_for(MAIN_URL_NAME)))
                error='Invalid id or password'
            return
        except TemplateNotFound:
                abort(404)
                 
    def logout():
        resp = session_logout(MAIN_URL_NAME, 'You were logged out')
        return resp
        
    def record_logging(path, ip, uid, action, data):
        with open(path, 'a+',encoding='utf-8') as fh:
            logging = datetime.today().strftime("%Y-%m-%d %H:%M:%S")+' (DB) '+action+' : '
            if uid:
                logging+=uid+' '
            logging += ip+' '+data+'\n'
            fh.write(logging)
            
    def get_db(data):
      """ key > THIS_COMPANY #str
          value['path'] > dv/data/db/this_company.db
          value['table_name'] > THIS_COMPANY
          item > 회사명, 프린트상호, 주소1 ...
          col[0] (주)두베스트, 직물 업체 ...
      """
      make_dict = {}
      temp_dict = {}
      temp_list = []
      for key, value in data.items():
        for item in value['column']:

          temp_gubun = []
          temp = query_db(value['path'], select_query(value['table_name'], item))
          for i, col in enumerate(temp): #tuple >> list
            temp_list.append(col[0])

          if ((str(key) is 'orderno') & (str(item) is 'OrderNo')):
            temp2 = query_db(value['path'], select_query(value['table_name'], value['column'][1]))

            for i, col in enumerate(temp2):  # tuple >> list
              temp_gubun.append(col[0])

          temp_dict[item] = [temp_list]

          if ((str(key) is 'orderno') & (str(item) is 'OrderNo')):
            temp_dict[item] = [temp_list, temp_gubun]
#            print(temp_dict[item])

          temp_list=[]
        make_dict[key] = temp_dict  
      return make_dict

    def init_get_db(data):
        orno = query_db(data['orderno']['path'], select_query(data['orderno']['table_name'],'계약일자'))
        orn = query_db(data['orderno']['path'], select_query(data['orderno']['table_name'],'OrderNo'))
        reg = re.compile(r'(?P<YYYY>\d{4})-(?P<MM>\d{2})')
        temp = {}
        tp_list = []
        for item in orno:
            if reg.search(item[0]):
                Y = reg.search(item[0])
                YYYY = Y.group("YYYY")
                MM = Y.group("MM")
                if MM not in tp_list:
                  tp_list.append(Y.group("MM"))
                temp[YYYY] = sorted(tp_list)
        
        return temp, orn

    
    def get_init_db():
        data = DB_CONT
        order_no = query_db(data['orderno']['path'], select_query(data['orderno']['table_name'],'OrderNo'))
        contract_date = query_db(data['orderno']['path'], select_query(data['orderno']['table_name'],'계약일자'))
        #print(contract_date[1][0])
        order_gubun = query_db(data['orderno']['path'], select_query(data['orderno']['table_name'],'OrderGubun'))
        init_d = []
        p =re.compile(r'(\d+-\d+-\d+)\s+')
        for i in range(0,len(order_no)):
            tm=''
            if p.search(contract_date[i][0]):
                tm = p.search(contract_date[i][0])
                tm =tm.group(0)
            init_d.append(order_no[i][0]+" / "+tm+" / "+order_gubun[i][0])
            
        return init_d

    def get_nego_db(checkcomp = 'DB'):
        data = DB_CONT
        Buyer = query_db(data['nego']['path'], select_query(data['nego']['table_name'], 'Buyer'))
        ConNo = query_db(data['nego']['path'], select_query(data['nego']['table_name'], 'ConNo'))
        BLNo = query_db(data['nego']['path'], select_query(data['nego']['table_name'], 'BLNo'))
        InvNo = query_db(data['nego']['path'], select_query(data['nego']['table_name'], 'InvNo'))
        LCNo = query_db(data['nego']['path'], select_query(data['nego']['table_name'], 'LCNo'))
        RefNo = query_db(data['nego']['path'], select_query(data['nego']['table_name'], 'RefNo'))

        nego_d = []


        for i in range(0, len(InvNo)):
            if checkcomp in InvNo[i][0]  :
                nego_d.append(str(InvNo[i][0]) + " / " + str(ConNo[i][0]) + " / " + str(BLNo[i][0]) + " / " + str(Buyer[i][0])+ " / " + str(LCNo[i][0]) + " / " + str(RefNo[i][0]))

        return nego_d

    def get_alldata_db(checkcomp = 'DB'):
        data = DB_CONT
        Buyer = query_db(data['alldata']['path'], select_query(data['alldata']['table_name'], 'buyer_order'))
        ConNo = query_db(data['alldata']['path'], select_query(data['alldata']['table_name'], 'tableSPL_ship'))        
        BLNo = query_db(data['alldata']['path'], select_query(data['alldata']['table_name'], 'ship_BL'))
        InvNo = query_db(data['alldata']['path'], select_query(data['alldata']['table_name'], 'OrderNo'))
        OrderGubun = query_db(data['alldata']['path'], select_query(data['alldata']['table_name'], 'OrderGubun'))
        order_LC = query_db(data['alldata']['path'], select_query(data['alldata']['table_name'], 'order_LC'))
        nego_ref = query_db(data['alldata']['path'], select_query(data['alldata']['table_name'], 'nego_ref'))

        nego_d = []

        for i in range(0, len(InvNo)):
            ConNoarr=ConNo[i][0].split(';')                    
            ConNo_ = ConNoarr[0]

            if checkcomp in InvNo[i][0]  :
                nego_d.append(InvNo[i][0]+"_"+OrderGubun[i][0] + " / " + ConNo_ + " / " + BLNo[i][0] + " / " + Buyer[i][0] + " / " + order_LC[i][0] + " / " + nego_ref[i][0])

        return nego_d

    def get_items_db(db_path, table_name, item):
        data = DB_CONT
        TERM = query_db(db_path, select_query(table_name, item))
        term_d = []

        for i in range(0, len(TERM)):
            if TERM[i][0] == 'nan' : 
                continue
            term_d.append(TERM[i][0])

        return term_d


    @app.route('/',methods=['POST','GET'])
    def main():
        try:
            error=None
            name=None
            mail=None
            adm = None
            AD=None
            CHECK_COMPANY = None
            cookie_id = request.cookies.get('cookie_id')
            qdb_cookie = query_db(USER_LIST,select_query('user_list','cookie_id'))
            q = query_db(USER_LIST,select_query('user_list'))
            
            if cookie_id:
                for i, cid in enumerate(qdb_cookie):
                   
                    if cookie_id in cid:
                        qdb_one=query_db(USER_LIST, select_query('user_list'),oneline=True, line_num=i)
                        update_session(qdb_one)
                        

            

            uid = session['name'] if 'uid' in session else None
            #print(uid)
            logins()
            if 'logged_in' not in session:
                print("no log")
#flash('Log in or Sign up')
            else:
                if not check_user_info():
                    resp = session_logout(MAIN_URL_NAME,'Your data has changed. Please login agian')
                    return resp
                else:
                    mail=session['mail']
                    name=session['name']
                    
                    if session['admin'] =='admin':
                        adm='ADMIN'
                    else:
                        adm='MEMBER'
            #print(adm)
            
                                                        
            #db_cont = get_db(DB_CONT)
            dbno,orn = init_get_db(DB_CONT)
            init_data = get_init_db()
            nego_data = get_nego_db()
            all_data =get_alldata_db()
            thiscompany_data         = get_items_db(DB_CONT['TERM']['path'], DB_CONT['TERM']['table_name'], "Term")
            agent_data               = get_items_db(DB_CONT['AGENT']['path'], DB_CONT['AGENT']['table_name'], "Agent")
            bank_data                = get_items_db(DB_CONT['BANK']['path'], DB_CONT['BANK']['table_name'], "은행")
            buyer_data               = get_items_db(DB_CONT['BUYER']['path'], DB_CONT['BUYER']['table_name'], "프린트Buyer")
            buyerbank_data           = get_items_db(DB_CONT['BUYERBANK']['path'], DB_CONT['BUYERBANK']['table_name'], "BankName")
            company_data             = get_items_db(DB_CONT['COMPANY']['path'], DB_CONT['COMPANY']['table_name'], "업체")
            contract_data            = get_items_db(DB_CONT['CONTRACT_CONDITION']['path'], DB_CONT['CONTRACT_CONDITION']['table_name'], "계약조건")
            currency_data            = get_items_db(DB_CONT['CURRENCY_UNIT']['path'], DB_CONT['CURRENCY_UNIT']['table_name'], "화폐단위")
            department_data          = get_items_db(DB_CONT['DEPARTMENT']['path'], DB_CONT['DEPARTMENT']['table_name'], "부서명")
            destination_data         = get_items_db(DB_CONT['DESTINATION']['path'], DB_CONT['DESTINATION']['table_name'], "프린트도착지1")
            exchangerate_data        = get_items_db(DB_CONT['EXCHANGE_RATE']['path'], DB_CONT['EXCHANGE_RATE']['table_name'], "환율")
            fairway_data             = get_items_db(DB_CONT['FAIRWAY']['path'], DB_CONT['FAIRWAY']['table_name'], "항로")
            forwarding_data          = get_items_db(DB_CONT['FORWARDING']['path'], DB_CONT['FORWARDING']['table_name'], "FORWARDING")
            convention_data          = get_items_db(DB_CONT['FTA_CONVENTION']['path'], DB_CONT['FTA_CONVENTION']['table_name'], "협정명칭")
            fta_item_number_data     = get_items_db(DB_CONT['FTA_ITEM_NUMBER']['path'], DB_CONT['FTA_ITEM_NUMBER']['table_name'], "품목번호")
            fta_origin_data          = get_items_db(DB_CONT['FTA_ORIGIN']['path'], DB_CONT['FTA_ORIGIN']['table_name'], "원산지")
            item_data                = get_items_db(DB_CONT['ITEM']['path'], DB_CONT['ITEM']['table_name'], "Item")
            luster_data              = get_items_db(DB_CONT['LUSTER']['path'], DB_CONT['LUSTER']['table_name'], "Luster")
            payment_data             = get_items_db(DB_CONT['PAYMENT']['path'], DB_CONT['PAYMENT']['table_name'], "Payment")
            purchasemethod_data      = get_items_db(DB_CONT['PURCHASE_METHOD']['path'], DB_CONT['PURCHASE_METHOD']['table_name'], "매입방식")
            quantity_data            = get_items_db(DB_CONT['QUANTITY']['path'], DB_CONT['QUANTITY']['table_name'], "수량단위")
            sea_data                 = get_items_db(DB_CONT['SEA']['path'], DB_CONT['SEA']['table_name'], '선적방법')
            shippingcompany_data     = get_items_db(DB_CONT['SHIPPING_COMPANY']['path'], DB_CONT['SHIPPING_COMPANY']['table_name'], "선박회사")
            term_data                = get_items_db(DB_CONT['TERM']['path'], DB_CONT['TERM']['table_name'], "Term")
            containeryard_data       = get_items_db(DB_CONT['CONTAINERYARD']['path'], DB_CONT['CONTAINERYARD']['table_name'], "Container")
            deliveryplace_data       = get_items_db(DB_CONT['DELIVERYPLACE']['path'], DB_CONT['DELIVERYPLACE']['table_name'], "Delivery_Place")
            localtrucking_data       = get_items_db(DB_CONT['LOCALTRUCKING']['path'], DB_CONT['LOCALTRUCKING']['table_name'], "Local_trucking")
            #menu
            contract_invoice_data = get_contract_invoice()
            inst_nego_bank = get_items_db(DB_CONT['inst_nego_bank']['path'],  DB_CONT['inst_nego_bank']['table_name'], "name")
			# company
            
            
            

            try:
                if session['company'] :
                    check_comp = session['company']
            except:
                check_comp = 'DB'
						
            return render_template(MAIN_URL, error=error,mail=mail, name=name, AD=AD, \
            adm = adm, db_cont='',db_YYYY=sorted(dbno, reverse=True),db_orn=orn, tplate=TEMPLATE, \
            init_data=init_data, nego_data=nego_data,
            thiscompany_data = thiscompany_data,
            agent_data = agent_data,
            bank_data = bank_data,
            buyer_data = buyer_data,
            buyerbank_data = buyerbank_data,
            company_data = company_data,
            contract_data = contract_data,
            currency_data = currency_data,
            department_data = department_data,
            destination_data = destination_data,
            exchangerate_data = exchangerate_data,
            fairway_data = fairway_data,
            forwarding_data = forwarding_data,
            convention_data = convention_data,
            fta_item_number_data = fta_item_number_data,
            fta_origin_data = fta_origin_data,
            item_data = item_data,
            luster_data = luster_data,
            payment_data = payment_data,
            purchasemethod_data = purchasemethod_data,
            quantity_data = quantity_data,
            sea_data =  sea_data,
            shippingcompany_data = shippingcompany_data,
            term_data = term_data,
            all_data = all_data,
            contract_invoice_data = contract_invoice_data,
            containeryard_data = containeryard_data,
            deliveryplace_data = deliveryplace_data,
            localtrucking_data = localtrucking_data,
            inst_nego_bank = inst_nego_bank,
            check_comp = check_comp
            )

        except TemplateNotFound:
            abort(404)

    @app.route('/logout/')
    def logout():
        try:
            record_logging(LOG_FILE, request.remote_addr, session['name']+' '+session['mail'], 'OUT','log out')
            resp=session_logout(MAIN_URL_NAME, 'You were logged out')
            return resp
        except TemplateNotFound:
            abort(404)
     
    @app.route('/sign_up/', methods=['POST','GET'])
    def sign_up():
      try:
        error=None
        if request.method == 'POST':
          form_data = request.json
          pp(form_data)
          app.config['SIGNUP'] = True
          name = request.json['name']
          #mail = request.json['mail']
          mail = ''
          #pp(mail)
          qdb=query_db(USER_LIST,select_query('user_list'))
          users_id = [ row[1] for row in qdb ]
          pp(users_id)
          if name in users_id:
            app.config['SIGNUP'] = False
            error = 'The ID already exists'
            return jsonify({'msg':error})
#return render_template(LAYOUT_URL, error=error)
          request_form = [name, request.json['password'], mail, 'client']
          request_form.append(app.config['NOW'])
          insert_vars = '(name, password, mail,authority, creation_date)' #creation_date
          update_db(USER_LIST, insert_query('user_list', insert_vars, 5), request_form)
          record_logging(LOG_FILE, request.remote_addr, name+' '+mail, 'SIGNUP','sign up'+name+' '+mail)
          return jsonify({"comp":"completed"})
#return redirect(url_for(MAIN_URL_NAME))
              
        return render_template(LAYOUT_URL, error=error)
      except TemplateNotFound:
        abort(404)
 
    @app.route('/login_r/', methods=['POST','GET'])
    def login_r():
      try:
        error=None
        if request.method == 'POST':
          form_data = request.json
          pp(form_data)
          name = request.json['name']
          #mail = request.json['mail']
          mail = ''
          pw = request.json['password']
          pp(mail)
          qdb=query_db(USER_LIST,select_query('user_list'))
          users_id = [ row[1] for row in qdb ]
          pp(qdb)
          for row in qdb:
              #if name == row[1] and mail == row[2] and pw == row[3]:
              if name == row[1]  and pw == row[3]: # remove email
                  record_logging(LOG_FILE, request.remote_addr, name+' '+mail, 'IN','logging')
                  update_session(row)
                  resp = make_response(redirect(url_for(MAIN_URL_NAME)))
                  return jsonify({"comp":"comp"})
                  if request.form.get('remember'):
                      cookie_id=set_random_id(20)
                      resp.set_cookie('cookie_id', cookie_id, max_age=app.config['REMEMBER_PERIOD'])
                      update_db(USER_LIST, update_query('user_list',['cookie_id'],'num'),[cookie_id,row[0]])
                      flash('You were logged in')
                      return make_response(redirect(url_for(MAIN_URL_NAME)))
          return jsonify({"msg":"error"})
#return redirect(url_for(MAIN_URL_NAME))
              
        return render_template(LAYOUT_URL, error=error)
      except TemplateNotFound:
        abort(404)




    @app.route('/insert_db/', methods=['POST','GET'])
    def insert_db():
      
      form_data = request.json
      insert_values=[]
      insert_values.append(form_data[1]['value'])
      insert_values.append(form_data[2]['value'])
      insert_values.append(form_data[3]['value'])
      insert_values.append(form_data[4]['value'])
      insert_values.append(form_data[5]['value'])
      insert_values.append(form_data[0]['value'])
      data = DB_CONT    
      ##### Add new data to a DB
      DB_NM = 'THIS_COMPANY'
      db_path = data[DB_NM]['path']
      table_name = data[DB_NM]['table_name']
      col_num = 6
      column_list = data[DB_NM]['column']
      col_list = ''
      for item in column_list:
        col_list+='"'+item+'",'
      col_list = col_list[:-1]
      
      update_db(db_path, insert_query(table_name,'('+col_list+')',col_num), insert_values)
      return jsonify({'msg':'complete'})
      ######
      
    
    @app.route('/get_prod/', methods=['POST','GET'])
    def get_prod():
      data = DB_CONT
      form_data = request.json
      YYYY = form_data[0]
      MM = form_data[1]
      order_no = query_db(data['orderno']['path'], select_query(data['orderno']['table_name'],'OrderNo'))
      contract_date = query_db(data['orderno']['path'], select_query(data['orderno']['table_name'],'계약일자'))
      reg = re.compile(r'(?P<YYYY>'+YYYY+')-(?P<MM>\d{2})')
      temp = {}
      tp_list = []
      i =0
      for item in contract_date:
          if reg.search(item[0]):
              #tp_list.append(item[0])
              tp_list.append(order_no[i][0])
          i +=1
      pp(tp_list)
      
    
      #if form_data == 'A':
      #  db_cont = ["Q","W","E","R"]
      #elif form_data == 'B':  
      #  db_cont = ["1","2","3"]
      #db_cont = get_db(DB_CONT)
      #db_cont = ["z","x","c"]
      return jsonify({'order_no':tp_list})

#    @app.context_processor
#    def utility_functions():
#        def print_in_console(message):
#              print (str(message))
#
#        return dict(mdebug=print_in_console)
      
    @app.route('/get_order_gubun/', methods=['POST','GET'])
    def get_order_gubun():
      data = DB_CONT
      form_data = request.json
      form_order_no = form_data
      #print(form_order_no)
      where_var = 'OrderNo="'+form_order_no+'"'
      order_no = query_db(data['orderno']['path'], selection_query(data['orderno']['table_name'],'OrderGubun',where_var))
      order_gubun =[]
      for item in order_no:
        order_gubun.append(item[0])
      pp(order_gubun)
      return jsonify({'order_gubun':order_gubun})
      
      
    @app.route('/get_order_buyer/', methods=['POST','GET'])
    def get_order_buyer():
      data = DB_CONT
      form_data = request.json
      form_order_no = form_data[0]
      form_order_gubun = form_data[1]
      #print(form_order_no)
      #print(form_order_gubun)
      form_order_gubun = form_order_gubun.replace(' ','')
      where_var = 'OrderNo="'+form_order_no+'"'
      where_var2 = 'OrderGubun=" '+form_order_gubun+'"'
      # ================================================================================
      #buyer
      order_no = query_db(data['orderno']['path'], selection_query_two(data['orderno']['table_name'],'Buyer',where_var,where_var2))
      # agnet
      order_no_agent = query_db(data['orderno']['path'], selection_query_two(data['orderno']['table_name'],'Agent',where_var,where_var2))
      # item
      order_no_item = query_db(data['orderno']['path'], selection_query_two(data['orderno']['table_name'],'Item',where_var,where_var2))
      # DN
      order_no_DN = query_db(data['orderno']['path'],selection_query_two(data['orderno']['table_name'], 'DN', where_var, where_var2))
      # LF
      order_no_LF = query_db(data['orderno']['path'],selection_query_two(data['orderno']['table_name'], 'LF', where_var, where_var2))
      # LS
      order_no_LS = query_db(data['orderno']['path'],selection_query_two(data['orderno']['table_name'], 'LS', where_var, where_var2))
      # remakr
      order_no_remark = query_db(data['orderno']['path'],selection_query_two(data['orderno']['table_name'], 'Remark', where_var, where_var2))
      # qty
      order_no_qty= query_db(data['orderno']['path'],selection_query_two(data['orderno']['table_name'], 'QTY', where_var, where_var2))
      # qty unit
      order_no_unit = query_db(data['orderno']['path'],selection_query_two(data['orderno']['table_name'], '수량단위', where_var, where_var2))
      # currency
      order_no_currency = query_db(data['orderno']['path'],selection_query_two(data['orderno']['table_name'], 'ORDER화폐단위', where_var, where_var2))
      # PRC
      order_no_PRC= query_db(data['orderno']['path'],selection_query_two(data['orderno']['table_name'], 'U_PRX', where_var,where_var2))
      # amount
      order_no_amount = query_db(data['orderno']['path'],selection_query_two(data['orderno']['table_name'], 'AMOUNT', where_var,where_var2))
      # ================================================================================

      # ================================================================================
      #buyer all data 
      order_no_other = query_db(data['BUYER']['path'], select_query(data['BUYER']['table_name'], 'Buyer'))
      #agent all data
      order_no_all_agent = query_db(data['AGENT']['path'], select_query(data['AGENT']['table_name'], 'Agent'))
      #item all data
      order_no_all_item = query_db(data['ITEM']['path'], select_query(data['ITEM']['table_name'], 'Item'))
      #DN all data
      #no data
      # LF all data
      # no data
      # LS all data
      order_no_all_LS = query_db(data['LUSTER']['path'], select_query(data['LUSTER']['table_name'], 'Luster'))
      # remark all data
      # no data
      # qty all data
      # no data
      # unit all data
      order_no_all_unit = query_db(data['QUANTITY']['path'], select_query(data['QUANTITY']['table_name'], '수량단위'))
      # currency all data
      order_no_all_currency = query_db(data['CURRENCY_UNIT']['path'], select_query(data['CURRENCY_UNIT']['table_name'], '화폐단위'))
      # PRC all data
      # no data
      # amount all data
      # no data
      # ================================================================================
                  
      #print('order_no_othter', order_no_other)
      order_buyer = []
      order_agent = []
      order_item = []
      order_DN = []
      order_LF = []
      order_LS = []
      order_remark = []
      order_qty = []
      order_unit = []
      order_currency = []
      order_PRC = []
      order_amount = []
      
      for item in order_no:
        order_buyer.append(item[0])
      for item in order_no_agent:
        order_agent.append(item[0])
      for item in order_no_item:
        order_item.append(item[0])
      for item in order_no_DN:
        order_DN.append(item[0])
      for item in order_no_LF:
        order_LF.append(item[0])
      for item in order_no_LS:
        order_LS.append(item[0])
      for item in order_no_remark:
        order_remark.append(item[0])
      for item in order_no_qty:
        order_qty.append(item[0])
      for item in order_no_unit:
        order_unit.append(item[0])
      for item in order_no_currency:
        order_unit.append(item[0])
      for item in order_no_PRC:
        order_PRC.append(item[0])
      for item in order_no_amount:
        order_amount.append(item[0])
        
      for item in order_no_other:
        if (item != 'nan') | (item != '') :
            order_buyer.append(item[0])
      for item in order_no_all_agent:
        if (item != 'nan') | (item != '') :
            order_agent.append(item[0])
      for item in order_no_all_item:
        if (item != 'nan') | (item != '') :
            order_item.append(item[0])
      for item in order_no_all_LS:
        if (item != 'nan') | (item != ''):
            order_LS.append(item[0])
      for item in order_no_all_unit:
        if (item != 'nan') | (item != ''):
            order_unit.append(item[0])
      for item in order_no_all_currency:
        if (item != 'nan') | (item != ''):
            order_currency.append(item[0])
            
      #pp(order_buyer)

      return jsonify({'order_buyer':order_buyer, 'order_agent':order_agent,'order_item':order_item,'order_DN':order_DN,'order_LF':order_LF,'order_LS':order_LS, 'order_remark':order_remark
                      ,'order_qty':order_qty,'order_unit':order_unit,'order_currency':order_currency,'order_PRC':order_PRC
                      ,'order_amount':order_amount})
    
    @app.route('/get_nego_DB/', methods=['POST','GET'])
    def get_nego_DB():
      data = DB_CONT
      form_data = request.json

      form_order_no = form_data
      #form_order_no = form_order_no[1:]

      #print("form_order_no:",form_order_no)
      
      #2020.01.29 When adding new data
      #with open('newdb.list','r') as fh:
          #for item in fh:
              #item = item.strip('\n')
              #if str(form_order_no).find(str(item)) != -1:
      form_order_no = form_order_no.replace(' ','')
      if str(form_order_no).find(str('_')) != -1:
        item = str(form_order_no)
        
        where_var = 'OrderNo="'+item.split('_')[0]+'"'
        where_var2 = 'OrderGubun="' + item.split('_')[1] + '"'
        path = 'dv/data/db/alldata.db'
        table_name = 'ALLDATA'
        select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))                  

        #print("where_var : ",where_var)
        #print("where_var2 : ",where_var2)
        #print("select_new : ",select_new)

        return jsonify({ 'check_new':'checked', 'select_new':select_new})

      #print( "check form_order_no ", str(form_order_no) )
      
      #form_order_no = str(form_order_no)

      #import re      
      #returnstr =re.findall(r'\d+', str(form_order_no))
      #print("returnstr",returnstr[0])
      
      # ================================================================================
      where_var = 'InvNo="'+form_order_no+'"'
      order_InvNo = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'InvNo',where_var))      
      order_Buyer = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Buyer',where_var))
      order_Agent = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Agent',where_var))
      nego_DueDate = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'DueDate',where_var))
      order_LC = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'LCNo',where_var))
      order_Item = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Item',where_var))
      order_DN = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'DN',where_var))
      order_LF = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'LF',where_var))
      order_LS = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'LS',where_var))
      order_remark = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Remark',where_var))
      order_Maker = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Maker',where_var))
      order_quantity = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Nego수량',where_var))
      order_dollor = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Nego달러',where_var))
      nego_won = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Nego원화',where_var))
      ship_name = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'선명',where_var))
      ship_docu_date = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'서류일자',where_var))
      ship_ship_date = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'선적일자',where_var))
      local_juk = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'적화보험료달러',where_var))
      local_juk2 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'적화보험료원화',where_var))
      local_sudollor = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'수출보험료달러',where_var))
      local_suwon = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'수출보험료원화',where_var))
      local_udollor = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'우편료달러',where_var))
      local_uwon = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'우편료원화',where_var))
      local_comm = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'기타달러',where_var))
      local_comm2 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'기타원화',where_var))
      local_actual = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Margin달러',where_var))
      local_actual2 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Margin원화',where_var))

      nego_m1 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'면장번호',where_var))
      nego_m2 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'면장신고일자',where_var))
      nego_m3 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'면장FOB달러',where_var))
      nego_m4 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'면장FOB원화',where_var))
      nego_date = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Nego일자',where_var))
      nego_ref = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'RefNo',where_var))
      nego_unit = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'화폐단위',where_var))
      nego_etc = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Nego기타통화',where_var))

      order_Payment = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'OrderPayment',where_var))
      ship_container = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'ConNo',where_var))
      ship_BL = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'BLNo',where_var))

      order_prc_unit = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'화폐단위_1',where_var))
      order_prc = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'NEGO단가',where_var))
      order_table_fob = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'FOB단가',where_var))
      order_term = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Term',where_var))
      nego_bank = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Bank',where_var))
      nego_invoice = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'InvoiceNo',where_var))
      order_arrive = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'도착지',where_var))
      order_com1 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Com1',where_var))
      order_com1_amount = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Amount1',where_var))
      order_com1_agent = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Agent_1',where_var))
      order_com2 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Com2',where_var))
      order_com2_amount = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Amount2',where_var))
      order_com2_agent = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Agent_2',where_var))
      order_com3 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Com3',where_var))
      order_com3_amount = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Amount3',where_var))
      order_com3_agent = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Agent_3',where_var))
      order_com4 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Com4',where_var))
      order_com4_amount = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Amount4',where_var))
      order_com4_agent = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Agent_4',where_var))
      nego_hwanga = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'환가요율',where_var))
      order_before_order = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'old_orderno',where_var))
      ship_chulgo = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'출고일',where_var))
      order_expect_price = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'Order예상선임',where_var))

      ship_lotNo = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'LoteNo',where_var))

      nego_deposit_date1 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'],'입금일자1',where_var))
      nego_deposit_amount1 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '입금액1', where_var))
      nego_deposit_date2 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '입금일자2', where_var))
      nego_deposit_amount2 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '입금액2', where_var))
      nego_deposit_date3 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '입금일자3', where_var))
      nego_deposit_amount3 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '입금액3', where_var))
      nego_deposit_date4 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '입금일자4', where_var))
      nego_deposit_amount4 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '입금액4', where_var))
      nego_deposit_date5 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '입금일자5', where_var))
      nego_deposit_amount5 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '입금액5', where_var))

      # Less ~ 계산서
      LessDeleyDollor = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], 'LessDeley달러', where_var))
      LessDeleyWon = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], 'LessDeley원화', where_var))
      LessDollor = query_db(data['nego']['path'],selection_query(data['nego']['table_name'], 'Less달러', where_var))
      LessWon = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], 'Less원화', where_var))
      DelayedAcqDollor = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '인수지연달러', where_var))
      DelayedAcqWon = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '인수지연원화', where_var))
      DateInvoice1 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '계산서일자1', where_var))
      BillInvoice1 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '계산서수량1', where_var))
      BillMoneyInvoice1 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '계산서금액1', where_var))
      BillDollorInvoice1 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '계산서금액달러1', where_var))
      DateInvoice2 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '계산서일자2', where_var))
      BillInvoice2 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '계산서수량2', where_var))
      BillMoneyInvoice2 = query_db(data['nego']['path'],selection_query(data['nego']['table_name'], '계산서금액2', where_var))
      BillDollorInvoice2 = query_db(data['nego']['path'],selection_query(data['nego']['table_name'], '계산서금액달러2', where_var))
      DateInvoice3 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '계산서일자3', where_var))
      BillInvoice3 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '계산서수량3', where_var))
      BillMoneyInvoice3 = query_db(data['nego']['path'],selection_query(data['nego']['table_name'], '계산서금액3', where_var))
      BillDollorInvoice3 = query_db(data['nego']['path'],selection_query(data['nego']['table_name'], '계산서금액달러3', where_var))
      DateInvoice4 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '계산서일자4', where_var))
      BillInvoice4 = query_db(data['nego']['path'], selection_query(data['nego']['table_name'], '계산서수량4', where_var))
      BillMoneyInvoice4 = query_db(data['nego']['path'],selection_query(data['nego']['table_name'], '계산서금액4', where_var))
      BillDollorInvoice4 = query_db(data['nego']['path'],selection_query(data['nego']['table_name'], '계산서금액달러4', where_var))

      #print("old data where_var", where_var)
      #print("old data order_InvNo", order_InvNo)

      return jsonify({'order_InvNo':order_InvNo,'order_Buyer':order_Buyer,'order_Agent':order_Agent, 'nego_DueDate':nego_DueDate,'order_LC':order_LC, 'order_Item': order_Item,
                      'order_DN': order_DN, 'order_LF': order_LF,'order_LS': order_LS,
                      'order_remark': order_remark, 'order_Maker': order_Maker, 'order_quantity':order_quantity,
                      'order_dollor':order_dollor,'nego_won':nego_won, 'ship_name':ship_name,
                      'ship_docu_date':ship_docu_date,'ship_ship_date':ship_ship_date, 'local_juk':local_juk,
                      'local_juk2':local_juk2,'local_sudollor':local_sudollor,'local_suwon':local_suwon,
                      'local_udollor':local_udollor,'local_uwon':local_uwon, 'local_comm':local_comm,
                      'local_comm2':local_comm2,'local_actual':local_actual,'local_actual2':local_actual2 ,
                      'nego_m1':nego_m1,'nego_m2':nego_m2,'nego_m3':nego_m3,'nego_m4':nego_m4,
                      'nego_date':nego_date,'nego_ref':nego_ref,'nego_unit':nego_unit,'nego_etc':nego_etc,
                      'order_Payment':order_Payment, 'ship_container':ship_container, 'ship_BL':ship_BL,
                      'order_prc_unit':order_prc_unit,'order_prc':order_prc,'order_table_fob':order_table_fob,
                      'order_term':order_term, 'nego_bank':nego_bank,'nego_invoice':nego_invoice,
                      'order_arrive':order_arrive,'order_com1':order_com1,'order_com1_amount':order_com1_amount,
                      'order_com1_agent':order_com1_agent,'order_com2':order_com2,'order_com2_amount':order_com2_amount,
                      'order_com2_agent':order_com2_agent,'order_com3':order_com3,'order_com3_amount':order_com3_amount,
                      'order_com3_agent':order_com3_agent,'order_com4':order_com4,'order_com4_amount':order_com4_amount,
                      'order_com4_agent':order_com4_agent,'nego_hwanga':nego_hwanga,'order_before_order':order_before_order,
                      'ship_chulgo':ship_chulgo,'order_expect_price':order_expect_price,'ship_lotNo':ship_lotNo,
                      'nego_deposit_date1':nego_deposit_date1, 'nego_deposit_amount1':nego_deposit_amount1,
                      'nego_deposit_date2': nego_deposit_date2, 'nego_deposit_amount2': nego_deposit_amount2,
                      'nego_deposit_date3': nego_deposit_date3, 'nego_deposit_amount3': nego_deposit_amount3,
                      'nego_deposit_date4': nego_deposit_date4, 'nego_deposit_amount4': nego_deposit_amount4,
                      'nego_deposit_date5': nego_deposit_date5, 'nego_deposit_amount5': nego_deposit_amount5,
                      'LessDeleyDollor':LessDeleyDollor,'LessDeleyWon':LessDeleyWon,'LessDollor':LessDollor,
                      'LessWon':LessWon,'DelayedAcqDollor':DelayedAcqDollor,'DelayedAcqWon':DelayedAcqWon,
                      'DateInvoice1':DateInvoice1,'BillInvoice1':BillInvoice1,'BillMoneyInvoice1':BillMoneyInvoice1,'BillDollorInvoice1':BillDollorInvoice1,
                      'DateInvoice2': DateInvoice2, 'BillInvoice2': BillInvoice2,'BillMoneyInvoice2': BillMoneyInvoice2, 'BillDollorInvoice2': BillDollorInvoice2,
                      'DateInvoice3': DateInvoice3, 'BillInvoice3': BillInvoice3,'BillMoneyInvoice3': BillMoneyInvoice3, 'BillDollorInvoice3': BillDollorInvoice3,
                      'DateInvoice4': DateInvoice4, 'BillInvoice4': BillInvoice4,'BillMoneyInvoice4': BillMoneyInvoice4, 'BillDollorInvoice4': BillDollorInvoice4
                      })
	
    @app.route('/get_contract_DB/', methods=['POST', 'GET'])
    def get_contract_DB():
        data = DB_CONT
        form_data = request.json

        form_order_no = form_data        
        form_order_no = form_order_no.replace(' ','')
        no = form_order_no.split('_')[0]
        gubun = form_order_no.split('_')[1]

        exist = 'none' # new , old
        item123 =[]

        # contract2에 DB가 있는지 본다 
        
        if str(form_order_no).find(str('_')) != -1:
            item = str(form_order_no)


            
            where_var = 'OrderNo="'+item.split('_')[0]+'"'
            where_var2 = 'OrderGubun="' + item.split('_')[1] + '"'
            path = 'dv/data/db/alldata.db'
            table_name = 'ALLDATA'

            #iCnt = query_db(data['alldata']['path'], selection_query(data['alldata']['table_name'], 'count(*)', where_var))[0][0]
            select_new2_list =query_db(data['alldata']['path'], selection_query(data['alldata']['table_name'], '*', where_var))

            #iCnt = query_db(data['alldata_contract2']['path'], selection_query(data['alldata_contract2']['table_name'], 'count(*)', where_var))[0][0]
            path_contract2 = 'dv/data/db/alldata_contract2.db'
            where1 = 'ordercon="'+item.split('_')[0]+'"'
            where2 = 'gubun="' + item.split('_')[1] + '"'
            contract2_found_db = query_db(path_contract2, selection_query_two('alldata_contract2','*',where1,where2))

            #print('contract2_found_db',contract2_found_db)
            if len(contract2_found_db) != 0 :
                exist = 'old'
                contract2_found_db = contract2_found_db[0]

            else :
                exist = 'new'
            
                for select_new in select_new2_list :

                    #select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))                  
                    #print('select_new',select_new)
                    # buyer
                    buyerselected = select_new[1]
                    #print('buyerselected',buyerselected)
                    path_buyer = data['BUYER']['path']
                    table_buyer = 'BUYER'
                    where_buyer = '프린트Buyer="'+str(buyerselected)+'"'
                    select_buyer = query_db(path_buyer,selection_query(table_buyer,'*',where_buyer))
                    #print('select_buyer',select_buyer)
                    # end buyer

                    # port 
                    portselected = select_new[150]
                    # end port

                    # term
                    termdata = select_new[8]
                    # end term
                    
                    # date order
                    datepicker = select_new[5]
                    # end date order 
                    
                    # date delivery 
                    datedelivery = select_new[7]
                    # end date of delivery
                    
                    # port of discharging
                    portofdisch = select_new[9]
                    # end port of discharging
                    
                    # payment_order
                    datapayment = select_new[10]
                    
                    # order_term
                    dataorderterm = select_new[11]
                    
                    # document_order
                    datadocument = select_new[13]

                    # item
                    data_item1 = select_new[27]
                    path_item = data['ITEM']['path']
                    table_item = 'ITEM'
                    where_item = 'Item="'+str(data_item1)+'"'
                    select_item = query_db(path_item,selection_query(table_item,'*',where_item))

                    # DN
                    dataDN = select_new[28]
                    
                    # LF
                    dataLF = select_new[29]

                    # L/S
                    dataLS = select_new[30]                    
                    path_luster = data['LUSTER']['path']
                    table_luster = 'LUSTER'
                    where_luster = 'Luster="'+str(dataLS)+'"'
                    select_luster = query_db(path_luster,selection_query(table_luster,'*',where_luster))
                    # end L/S

                    # qty  order_quantity  34
                    dataQTY = select_new[34]
                    # prc order_prc  37
                    dataPRC = select_new[37]

                    #amount
                    dataAmount = select_new[68]

                    # QTY_order 35
                    QTYorder = select_new[35]

                    # PRC_order 36
                    PRCorder = select_new[36]
                    #print('QTYorder:', QTYorder)
                    #print('item check', )
                    if len(select_item) == 0:
                        select_item = ['']
                    item1 = select_item
                    item1.append(dataDN)
                    item1.append(dataLF)
                    item1.append(dataQTY)
                    item1.append(dataPRC)
                    item1.append(dataAmount)
                    item1.append(QTYorder)
                    item1.append(PRCorder)
                    item123.append(item1)    
         

        if exist == 'new' :
            #print('exist' , exist)
            return jsonify({'check_new':'checked','buyer':select_buyer, 'no':no, 'gubun':gubun
                        ,'port':portselected,'term':termdata, 'deliverydate':datedelivery
                        ,'orderdate':datepicker,'portofdisch':portofdisch,'datapayment':datapayment
                        ,'dataorderterm':dataorderterm,'datadocument':datadocument,
                        'item123':item123,'luster':select_luster})
        else :
            #print('exist' , exist)
            return jsonify({
                'check_new':'notchecked',
                'iddb' : contract2_found_db[0],
                'idgb' :contract2_found_db[1] ,
                'contract2_invoice_shipper': contract2_found_db[2],
                'date_contract2_ship'  : contract2_found_db[3],
                'contract2_loading' : contract2_found_db[30],
                'contract2_insurance' : contract2_found_db[5],
                'date_contract2_contract_date' : contract2_found_db[7],
                'contract2_no' : contract2_found_db[8],
                'contract2_reference' : contract2_found_db[9],
                'contract2_discharging' : contract2_found_db[10],
                'contract2_trans' : contract2_found_db[11],
                'contract2_shipment' : contract2_found_db[12],
                'contract2_payment' : contract2_found_db[13],
                'contract2_remark' : contract2_found_db[6],
                'contract2_term'  : contract2_found_db[29],
                'contract2_goods1' : contract2_found_db[15],
                'contract2_goods2' : contract2_found_db[19],
                'contract2_goods3' : contract2_found_db[23],
                'contract2_quantity_1' : contract2_found_db[16],
                'contract2_quantity_2' : contract2_found_db[20],
                'contract2_quantity_3' : contract2_found_db[24],
                'contract2_quantity_total' : contract2_found_db[27],
                'contract2_price_1' : contract2_found_db[17],
                'contract2_price_2' : contract2_found_db[21],
                'contract2_price_3' : contract2_found_db[24],
                'contract2_amount_1' : contract2_found_db[18],
                'contract2_amount_2' : contract2_found_db[22],
                'contract2_amount_3' : contract2_found_db[26],
                'contract2_amount_total' : contract2_found_db[28]})


        msg = '[error] error'
        return jsonify({'msg':msg})

    @app.route('/get_shippingdocs_DB/', methods=['POST', 'GET'])
    def get_shippingdocs_DB():
        data = DB_CONT
        form_data = request.json

        form_order_no = form_data
        #form_order_no = form_order_no[1:]

        form_order_no = form_order_no.replace(' ','')
        no = form_order_no.split('_')[0]
        gubun = form_order_no.split('_')[1]

        exist = 'none' # new , old
        item123 =[]
        #with open('newdb.list','r') as fh:
        #    for item in fh:
        #        item = item.strip('\n')
                
        #        if str(item).find(str(no)) != -1:
        if str(form_order_no).find(str('_')) != -1:
            item = str(form_order_no)


            exist = 'new'
            
            where_var = 'OrderNo="'+item.split('_')[0]+'"'
            where_var2 = 'OrderGubun="' + item.split('_')[1] + '"'
            path = 'dv/data/db/alldata.db'
            table_name = 'ALLDATA'

            select_new2_list =query_db(data['alldata']['path'], selection_query(data['alldata']['table_name'], '*', where_var))

            for select_new in select_new2_list : 
                #select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))                  

                # buyer
                buyerselected = select_new[1]
                path_buyer = data['BUYER']['path']
                table_buyer = 'BUYER'
                where_buyer = '프린트Buyer="'+str(buyerselected)+'"'
                select_buyer = query_db(path_buyer,selection_query(table_buyer,'*',where_buyer))
                # end buyer

                # inv_ship
                inv_ship = select_new[53]
                # end inv_ship

                # actual ship 
                actual_ship = select_new[54]

                # port 
                portselected = select_new[150]
                # end port

                # term
                termdata = select_new[8]
                # end term
                
                # date order
                datepicker = select_new[5]
                # end date order 
                
                # date delivery 
                datedelivery = select_new[7]
                # end date of delivery
                
                # port of discharging
                portofdisch = select_new[9]
                path_dest = data['DESTINATION']['path']
                table_dest = 'DESTINATION'
                where_dest = '프린트도착지1="'+str(portofdisch)+'"'
                select_dest = query_db(path_dest,selection_query(table_dest,'*',where_dest))
                # end port of discharging
                
                # payment_order
                datapayment = select_new[10]
                
                # order_term
                dataorderterm = select_new[11]
                
                # document_order
                datadocument = select_new[13]

                # ship name
                data_ship_name = select_new[58]

                # item
                data_item1 = select_new[27]
                path_item = data['ITEM']['path']
                table_item = 'ITEM'
                where_item = 'Item="'+str(data_item1)+'"'
                select_item = query_db(path_item,selection_query(table_item,'*',where_item))
                print('select_item', select_item)

                # DN
                dataDN = select_new[28]
                
                # LF
                dataLF = select_new[29]

                # L/S
                dataLS = select_new[30]                    
                path_luster = data['LUSTER']['path']
                table_luster = 'LUSTER'
                where_luster = 'Luster="'+str(dataLS)+'"'
                select_luster = query_db(path_luster,selection_query(table_luster,'*',where_luster))
                # end L/S

                # qty  order_quantity  34
                dataQTY = select_new[34]
                # prc order_prc  37
                dataPRC = select_new[37]

                #amount
                dataAmount = select_new[39]

                # QTY_order 35
                QTYorder = select_new[35]

                # PRC_order 36
                PRCorder = select_new[36]

                # user
                user_name = session['name']

                if len(select_item) == 0:
                    select_item = ['']
                item1 = select_item
                item1.append(dataDN)
                item1.append(dataLF)
                item1.append(dataQTY)
                item1.append(dataPRC)
                item1.append(dataAmount)
                item1.append(QTYorder)
                item1.append(PRCorder)
                item123.append(item1)            

                '''
                if item.split('_')[1] == 'A' :                        
                    item1 = select_item
                    item1.append(dataDN)
                    item1.append(dataLF)
                    item1.append(dataQTY)
                    item1.append(dataPRC)
                    item1.append(dataAmount)
                    item123.append(item1)                        
                elif item.split('_')[1] == 'B' :
                    item2 = select_item
                    item2.append(dataDN)
                    item2.append(dataLF)
                    item2.append(dataQTY)
                    item2.append(dataPRC)
                    item2.append(dataAmount)
                    item123.append(item2)
                elif item.split('_')[1] == 'C' :
                    item3 = select_item
                    item3.append(dataDN)
                    item3.append(dataLF)
                    item3.append(dataQTY)
                    item3.append(dataPRC)
                    item3.append(dataAmount)
                    item123.append(item3)
                '''
        if exist == 'new' :
            
            return jsonify({'check_new':'checked','buyer':select_buyer, 'no':no, 'gubun':gubun
                        ,'port':portselected,'term':termdata, 'deliverydate':datedelivery
                        ,'orderdate':datepicker,'portofdisch':portofdisch,'datapayment':datapayment
                        ,'dataorderterm':dataorderterm,'datadocument':datadocument,
                        'item123':item123, 'inv_ship':inv_ship,'select_dest':select_dest,
                        'data_ship_name':data_ship_name, 'actual_ship':actual_ship, 'luster':select_luster
                        ,'user_name':user_name})

        msg = '[error] error'
        return jsonify({'msg':msg})

    @app.route('/get_deliveryorder_DB/', methods=['POST', 'GET'])
    def get_deliveryorder_DB():
        data = DB_CONT
        form_data = request.json

        form_order_no = form_data
        #form_order_no = form_order_no[1:]

        #print('get_deliveryorder_DB', form_order_no)

        form_order_no = form_order_no.replace(' ','')
        no = form_order_no.split('_')[0]
        gubun = form_order_no.split('_')[1]

        exist = 'none' # new , old
        item123 =[]
        #with open('newdb.list','r') as fh:
        #    for item in fh:
        #        item = item.strip('\n')
                
        #        if str(item).find(str(no)) != -1:

        if str(form_order_no).find(str('_')) != -1:
            item = str(form_order_no)

            exist = 'new'
            
            where_var = 'OrderNo="'+item.split('_')[0]+'"'
            where_var2 = 'OrderGubun="' + item.split('_')[1] + '"'
            path = 'dv/data/db/alldata.db'
            table_name = 'ALLDATA'
            select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))                  

            #print('where_var', where_var)
            #print('where_var2', where_var2)
            #print('select_new[0]' , select_new)
            # user
            user_name = session['name']


            # buyer
            buyerselected = select_new[0][1]
            path_buyer = data['BUYER']['path']
            table_buyer = 'BUYER'
            where_buyer = '프린트Buyer="'+str(buyerselected)+'"'
            select_buyer = query_db(path_buyer,selection_query(table_buyer,'*',where_buyer))
            # end buyer

            # port of discharging
            portofdisch = select_new[0][9]
            path_dest = data['DESTINATION']['path']
            table_dest = 'DESTINATION'
            where_dest = '프린트도착지1="'+str(portofdisch)+'"'
            select_dest = query_db(path_dest,selection_query(table_dest,'*',where_dest))
            # end port of discharging
            
            # item
            data_item1 = select_new[0][27]
            path_item = data['ITEM']['path']
            table_item = 'ITEM'
            where_item = 'Item="'+str(data_item1)+'"'
            select_item = query_db(path_item,selection_query(table_item,'*',where_item))

            # DN
            dataDN = select_new[0][28]
            
            # LF
            dataLF = select_new[0][29]

            # L/S
            dataLS = select_new[0][30]                    
            path_luster = data['LUSTER']['path']
            table_luster = 'LUSTER'
            where_luster = 'Luster="'+str(dataLS)+'"'
            select_luster = query_db(path_luster,selection_query(table_luster,'*',where_luster))
            # end L/S

            # qty  order_quantity  34
            dataQTY = select_new[0][34]
            # prc order_prc  37
            dataPRC = select_new[0][37]

            #amount
            dataAmount = select_new[0][39]

            # supply name
            supplydata=select_new[0][44]                    
            supplydatasplit=supplydata.split(';')
            #print('supplydata'+supplydata)
            supplyname = supplydatasplit[0]
            
            if item.split('_')[1] == 'A' :                        
                item1 = select_item
                item1.append(dataDN)
                item1.append(dataLF)
                item1.append(dataQTY)
                item1.append(dataPRC)
                item1.append(dataAmount)
                item123.append(item1)                        
            elif item.split('_')[1] == 'B' :
                item2 = select_item
                item2.append(dataDN)
                item2.append(dataLF)
                item2.append(dataQTY)
                item2.append(dataPRC)
                item2.append(dataAmount)
                item123.append(item2)
            elif item.split('_')[1] == 'C' :
                item3 = select_item
                item3.append(dataDN)
                item3.append(dataLF)
                item3.append(dataQTY)
                item3.append(dataPRC)
                item3.append(dataAmount)
                item123.append(item3)

        if exist == 'new' :
            
            return jsonify({'check_new':'checked','buyer':select_buyer, 'no':no, 'gubun':gubun
                        ,'portofdisch':portofdisch,'item123':item123,'select_dest':select_dest,
                        'luster':select_luster,'user_name':user_name,'supplyname':supplyname})

        msg = '[error] error'
        return jsonify({'msg':msg})

    @app.route('/get_max_DBB/', methods=['POST', 'GET'])
    def get_max_DBB():
        data = DB_CONT

        #print('CHECK_COMPANY',CHECK_COMPANY)

        check_company = request.json
        #print(check_company)
        #print(check_company[0])
        #print(check_company[1])

        order_no = query_db(data['nego']['path'],
                            selection_query_order(data['nego']['table_name'], 'InvNo', 'InvNo'))
        order_DBB=[]
        order_gubun=[]
        #for item in order_no :
        #    if str(item).find('DBB') == -1 :
        #        continue
        #    print('item old', item)
        #    item1=str(item)[2:9]
        #    item2=str(item)[9:10]
        #    order_DBB.append(item1)
        #    order_gubun.append(item2)
        #    break

        dbdb_alldata = query_db(data['alldata']['path'],
                            selection_query_order(data['alldata']['table_name'], 'OrderNo', 'OrderNo'))

        order_DBB_all = []
        
        for item in dbdb_alldata:
            if str(item).find((check_company[0])) == -1:
                continue

            #print('item ',item)
            strcheck ='DB' + check_company[1]
            #print('strcheck ',strcheck)
            if str(item[0]).find(strcheck) :
                continue

            #print('check_company', CHECK_COMPANY)

            
            #item1 = str(item)[2:9]
            order_DBB_all.append(item[0])
            break

        checkNew = False

        if len(order_DBB_all) > 0 :
            order_DBB = order_DBB_all
            #order_DBB = str(check_company[0] + check_company[1] + order_DBB_all[0])
            order_gubun = 'A'
            checkNew = True
        else :
            order_DBB .append(str(check_company[0] + check_company[1] +'0000'))
            order_gubun = 'A'
            checkNew = True

        

        #print('order_gubun', order_gubun,'order_DBB',order_DBB, 'checkNew',checkNew)
        return jsonify({'order_gubun': order_gubun,'order_DBB':order_DBB, 'checkNew':checkNew})
        
        
#####2020.01.15 Code related to all DB retrieval
    def init_query_db_all(vDB, db_path, limit_cnt, item_name):
        import math
        final_db = {}
        total_count = query_db(db_path,select_query(item_name,'count(*)'))[0][0]
        list_count = math.ceil(int(total_count)/100)
        for item in vDB:
            final_db[item] = query_db(db_path,select_query(item_name+' limit '+limit_cnt, item))
        return final_db, vDB, total_count, list_count

    def query_db_all_sel(vDB,DB):
        import math
        final_data=[]
        final_db = {}
        vDB = ''
        view_db =''
        limit_cnt = DB["inx"].split('_')
        DB = limit_cnt[0].split(',')
        db_name = str(limit_cnt[1])
        if(db_name == 'alldata'):
            vDB = DB_CONT["alldata"]["view_db"]
            view_db = DB_CONT[db_name]['view_db']
        elif(db_name == 'BUYER'):
            view_db = DB_CONT["BUYER"]["column"]

        db_path=DB_CONT[db_name]['path']

        total_count = query_db(db_path,select_query(db_name,'count(*)'))[0][0]
        list_count = math.ceil(int(total_count)/100)
        for item in view_db:
            final_db[item] = query_db(db_path,select_query(db_name+' limit '+limit_cnt[0], item))
            try:
                del final_db[item][100:]
            except:
                pass
        return final_db,list_count, view_db, db_name

    @app.route('/add_item_db_buyer/', methods=['POST','GET'])
    def add_item_db_buyer():
        if request.method == 'POST':
            _code = request.form["_code"]
            _buyer= request.form["_buyer"]
            _prt_buyer = request.form["_prt_buyer"]
            _addr_1 = request.form["_addr_1"]
            _addr_2 = request.form["_addr_2"]
            _tel = request.form["_tel"]
            _fax = request.form["_fax"]
            _dam = request.form["_dam"]
            _jik = request.form["_jik"]
            _mail = request.form["_mail"]
            _mobile = request.form["_mobile"]
            _level = request.form["_level"]
            _bc = request.form["_bc"]
            _grade_date = request.form["_grade_date"]
            request_form = [_code, _buyer, _prt_buyer, _addr_1, _addr_2, _tel, _fax, _dam, _jik, _mail, _mobile, _level, _bc, _grade_date]
            insert_vars = '(코드, Buyer, 프린트Buyer, 주소1, 주소2, TEL, FAX, DAM, JIK, EMAIL, MOBIL, LEVEL, 바이어코드, 등급발행일)'
            update_db(DB_CONT['BUYER']['path'], insert_query('BUYER', insert_vars,14), request_form)
            return redirect(url_for('viewDB'))
        return

    @app.route('/viewDB/', methods=['POST','GET'])
    def viewDB():
        try:
            record_logging(LOG_FILE, request.remote_addr, session['name']+' '+session['mail'], 'VIEW','view DB')
            limit_cnt = 0
            final_data = []
            final_db ={}

            f = request.json

            if f:
                final_data, list_count, view_db, db_name = query_db_all_sel(DB_CONT["alldata"]["view_db"],f)
                return jsonify({'data':final_data,'list_count':list_count, 'view_db':view_db, 'db_name':db_name})
            else:
                limit_cnt = '0,100'

            final_db,db_data,total_count,data_list = init_query_db_all(DB_CONT["alldata"]["view_db"], DB_CONT['alldata']['path'], limit_cnt, 'alldata')
            final_db_buy,db_data_buy,total_count_buy,data_list_buy = init_query_db_all(DB_CONT["BUYER"]["column"], DB_CONT['BUYER']['path'], limit_cnt, 'BUYER')
            final_db_buyer_bank,db_data_buyer_bank,total_count_buyer_bank,data_list_buyer_bank = init_query_db_all(DB_CONT["BUYERBANK"]["column_v"], DB_CONT['BUYERBANK']['path'], '0,1000', 'BUYERBANK')
            f_agent,db_agent,tc_agent,list_agent = init_query_db_all(DB_CONT["AGENT"]["column_v"], DB_CONT['AGENT']['path'], '0,1000', 'AGENT')
            F_BANK,DB_BANK,TC_BANK,LIST_BANK = init_query_db_all(DB_CONT["BANK"]["column_v"], DB_CONT['BANK']['path'], '0,1000', 'BANK')

            F_COMPANY,DB_COMPANY,TC_COMPANY,LIST_COMPANY = init_query_db_all(DB_CONT["COMPANY"]["column_v"], DB_CONT['COMPANY']['path'], '0,1000', 'COMPANY')
            F_CONTAINERYARD,DB_CONTAINERYARD,TC_CONTAINERYARD,LIST_CONTAINERYARD = init_query_db_all(DB_CONT["CONTAINERYARD"]["column_v"], DB_CONT['CONTAINERYARD']['path'], '0,1000', 'CONTAINERYARD')
            F_CONTRACT_CONDITION,DB_CONTRACT_CONDITION,TC_CONTRACT_CONDITION,LIST_CONTRACT_CONDITION = init_query_db_all(DB_CONT["CONTRACT_CONDITION"]["column_v"], DB_CONT['CONTRACT_CONDITION']['path'], '0,1000', 'CONTRACT_CONDITION')
            F_CURRENCY_UNIT,DB_CURRENCY_UNIT,TC_CURRENCY_UNIT,LIST_CURRENCY_UNIT = init_query_db_all(DB_CONT["CURRENCY_UNIT"]["column_v"], DB_CONT['CURRENCY_UNIT']['path'], '0,1000', 'CURRENCY_UNIT')
            F_DELIVERYPLACE,DB_DELIVERYPLACE,TC_DELIVERYPLACE,LIST_DELIVERYPLACE = init_query_db_all(DB_CONT["DELIVERYPLACE"]["column_v"], DB_CONT['DELIVERYPLACE']['path'], '0,1000', 'DELIVERYPLACE')
            F_DEPARTMENT,DB_DEPARTMENT,TC_DEPARTMENT,LIST_DEPARTMENT = init_query_db_all(DB_CONT["DEPARTMENT"]["column_v"], DB_CONT['DEPARTMENT']['path'], '0,1000', 'DEPARTMENT')
            F_DESTINATION,DB_DESTINATION,TC_DESTINATION,LIST_DESTINATION = init_query_db_all(DB_CONT["DESTINATION"]["column_v"], DB_CONT['DESTINATION']['path'], '0,1000', 'DESTINATION')
            F_EXCHANGE_RATE,DB_EXCHANGE_RATE,TC_EXCHANGE_RATE,LIST_EXCHANGE_RATE = init_query_db_all(DB_CONT["EXCHANGE_RATE"]["column_v"], DB_CONT['EXCHANGE_RATE']['path'], '0,1000', 'EXCHANGE_RATE')
            F_FAIRWAY,DB_FAIRWAY,TC_FAIRWAY,LIST_FAIRWAY = init_query_db_all(DB_CONT["FAIRWAY"]["column_v"], DB_CONT['FAIRWAY']['path'], '0,1000', 'FAIRWAY')
            F_FORWARDING,DB_FORWARDING,TC_FORWARDING,LIST_FORWARDING = init_query_db_all(DB_CONT["FORWARDING"]["column_v"], DB_CONT['FORWARDING']['path'], '0,1000', 'FORWARDING')
            F_FTA_CONVENTION,DB_FTA_CONVENTION,TC_FTA_CONVENTION,LIST_FTA_CONVENTION = init_query_db_all(DB_CONT["FTA_CONVENTION"]["column_v"], DB_CONT['FTA_CONVENTION']['path'], '0,1000', 'FTA_CONVENTION')
            F_FTA_ITEM_NUMBER,DB_FTA_ITEM_NUMBER,TC_FTA_ITEM_NUMBER,LIST_FTA_ITEM_NUMBER = init_query_db_all(DB_CONT["FTA_ITEM_NUMBER"]["column_v"], DB_CONT['FTA_ITEM_NUMBER']['path'], '0,1000', 'FTA_ITEM_NUMBER')
            F_FTA_ORIGIN,DB_FTA_ORIGIN,TC_FTA_ORIGIN,LIST_FTA_ORIGIN = init_query_db_all(DB_CONT["FTA_ORIGIN"]["column_v"], DB_CONT['FTA_ORIGIN']['path'], '0,1000', 'FTA_ORIGIN')
            F_ITEM,DB_ITEM,TC_ITEM,LIST_ITEM = init_query_db_all(DB_CONT["ITEM"]["column_v"], DB_CONT['ITEM']['path'], '0,1000', 'ITEM')
            F_LOCALTRUCKING,DB_LOCALTRUCKING,TC_LOCALTRUCKING,LIST_LOCALTRUCKING = init_query_db_all(DB_CONT["LOCALTRUCKING"]["column_v"], DB_CONT['LOCALTRUCKING']['path'], '0,1000', 'LOCALTRUCKING')
            F_LUSTER,DB_LUSTER,TC_LUSTER,LIST_LUSTER = init_query_db_all(DB_CONT["LUSTER"]["column_v"], DB_CONT['LUSTER']['path'], '0,1000', 'LUSTER')
            F_PAYMENT,DB_PAYMENT,TC_PAYMENT,LIST_PAYMENT = init_query_db_all(DB_CONT["PAYMENT"]["column_v"], DB_CONT['PAYMENT']['path'], '0,1000', 'PAYMENT')
            F_PURCHASE_METHOD,DB_PURCHASE_METHOD,TC_PURCHASE_METHOD,LIST_PURCHASE_METHOD = init_query_db_all(DB_CONT["PURCHASE_METHOD"]["column_v"], DB_CONT['PURCHASE_METHOD']['path'], '0,1000', 'PURCHASE_METHOD')
            F_QUANTITY,DB_QUANTITY,TC_QUANTITY,LIST_QUANTITY = init_query_db_all(DB_CONT["QUANTITY"]["column_v"], DB_CONT['QUANTITY']['path'], '0,1000', 'QUANTITY')
            F_SEA,DB_SEA,TC_SEA,LIST_SEA = init_query_db_all(DB_CONT["SEA"]["column_v"], DB_CONT['SEA']['path'], '0,1000', 'SEA')
            F_SHIPPING_COMPANY,DB_SHIPPING_COMPANY,TC_SHIPPING_COMPANY,LIST_SHIPPING_COMPANY = init_query_db_all(DB_CONT["SHIPPING_COMPANY"]["column_v"], DB_CONT['SHIPPING_COMPANY']['path'], '0,1000', 'SHIPPING_COMPANY')
            F_TERM,DB_TERM,TC_TERM,LIST_TERM = init_query_db_all(DB_CONT["TERM"]["column_v"], DB_CONT['TERM']['path'], '0,1000', 'TERM')
                
            return render_template('viewdb.html', \
                                    data_list=data_list, total_count=total_count,\
                                    final_db=final_db, db_data=db_data,\
                                    final_db_buy=final_db_buy, db_data_buy=db_data_buy,\
                                    final_db_buyer_bank=final_db_buyer_bank, db_data_buyer_bank=db_data_buyer_bank,\
                                    f_agent=f_agent,db_agent=db_agent,\
                                    F_BANK=F_BANK,DB_BANK=DB_BANK,\
                                    F_COMPANY=F_COMPANY,DB_COMPANY=DB_COMPANY,\
                                    F_CONTAINERYARD=F_CONTAINERYARD,DB_CONTAINERYARD=DB_CONTAINERYARD,\
                                    F_CONTRACT_CONDITION=F_CONTRACT_CONDITION,DB_CONTRACT_CONDITION=DB_CONTRACT_CONDITION,\
                                    F_CURRENCY_UNIT=F_CURRENCY_UNIT,DB_CURRENCY_UNIT=DB_CURRENCY_UNIT,\
                                    F_DELIVERYPLACE=F_DELIVERYPLACE,DB_DELIVERYPLACE=DB_DELIVERYPLACE,\
                                    F_DEPARTMENT=F_DEPARTMENT,DB_DEPARTMENT=DB_DEPARTMENT,\
                                    F_DESTINATION=F_DESTINATION,DB_DESTINATION=DB_DESTINATION,\
                                    F_EXCHANGE_RATE=F_EXCHANGE_RATE,DB_EXCHANGE_RATE=DB_EXCHANGE_RATE,\
                                    F_FAIRWAY=F_FAIRWAY,DB_FAIRWAY=DB_FAIRWAY,\
                                    F_FORWARDING=F_FORWARDING,DB_FORWARDING=DB_FORWARDING,\
                                    F_FTA_CONVENTION=F_FTA_CONVENTION,DB_FTA_CONVENTION=DB_FTA_CONVENTION,\
                                    F_FTA_ITEM_NUMBER=F_FTA_ITEM_NUMBER,DB_FTA_ITEM_NUMBER=DB_FTA_ITEM_NUMBER,\
                                    F_FTA_ORIGIN=F_FTA_ORIGIN,DB_FTA_ORIGIN=DB_FTA_ORIGIN,\
                                    F_ITEM=F_ITEM,DB_ITEM=DB_ITEM,\
                                    F_LOCALTRUCKING=F_LOCALTRUCKING,DB_LOCALTRUCKING=DB_LOCALTRUCKING,\
                                    F_LUSTER=F_LUSTER,DB_LUSTER=DB_LUSTER,\
                                    F_PAYMENT=F_PAYMENT,DB_PAYMENT=DB_PAYMENT,\
                                    F_PURCHASE_METHOD=F_PURCHASE_METHOD,DB_PURCHASE_METHOD=DB_PURCHASE_METHOD,\
                                    F_QUANTITY=F_QUANTITY,DB_QUANTITY=DB_QUANTITY,\
                                    F_SEA=F_SEA,DB_SEA=DB_SEA,\
                                    F_SHIPPING_COMPANY=F_SHIPPING_COMPANY,DB_SHIPPING_COMPANY=DB_SHIPPING_COMPANY,\
                                    F_TERM=F_TERM,DB_TERM=DB_TERM,\
                                    total_count_buy=total_count_buy, data_list_buy=data_list_buy)
        except TemplateNotFound:
            abort(404)
  
    @app.route('/add_dt', methods=['POST','GET'])
    def add_dt():
        fd =  request.form
        in_vars = DB_CONT[request.form["db"]]["column"]
        pp(in_vars)
        temp = []
        temp_iv = '('
        for item in in_vars:
            temp.append(request.form[item])
            if item == in_vars[-1]:
                temp_iv+=item+')'
            else:
                temp_iv+=item+','
        pp(temp_iv)
        update_db(DB_CONT[request.form["db"]]['path'], insert_query(request.form["db"], temp_iv, len(in_vars)), temp)
        return make_response(redirect('viewDB'))

    @app.route('/edit_view', methods=['POST','GET'])
    def edit_view():
        fd = request.json
        pp(fd)
        update_db(DB_CONT[fd["_db"]]["path"], update_query(fd["_db"], DB_CONT[fd["_db"]]["column"], 'pk'), fd["_data"])
        return "complete"

    @app.route('/delete_view', methods=['POST','GET'])
    def delete_view():
        fd = request.json
        pp(fd)
        update_db(DB_CONT[fd["_db"]]["path"], delete_query(fd["_db"],'pk'), [fd["_pk"]])
        return "complete"



    @app.route('/modi_all', methods=['POST','GET'])
    def modi_all():
        form_data = request.json
        db_new = form_data['inx']
        print("==========form")
        #pp(db_new)
        import copy
        db_list = copy.deepcopy(DB_CONT["alldata"]["view_db"])
        del db_list[-1]
        print("====================alldata=====")
        #pp(DB_CONT["alldata"]["view_db"])
        update_db(DB_CONT["alldata"]["path"], update_query("alldata", db_list, 'pk'), db_new)
        
        return make_response(redirect('viewDB'))
  
             
    @app.route('/del_all_data', methods=['POST','GET'])
    def del_all_data():
        form_data = request.json
        inx = form_data['inx']
        update_db(DB_CONT["alldata"]["path"], delete_query("alldata",'pk'), [inx])
        
        return make_response(redirect('viewDB'))
         
                                            
    @app.route('/modi_buyer', methods=['POST','GET'])
    def modi_buyer():
        form_data = request.json
        db_new = form_data['inx']
        import copy
        db_list = copy.deepcopy(DB_CONT["BUYER"]["column"])
        del db_list[-1]
        #pp(DB_CONT["BUYER"]["column"])
        update_db(DB_CONT["BUYER"]["path"], update_query("BUYER", db_list, 'pk'), db_new)
        
        return make_response(redirect('viewDB'))
        
        
    @app.route('/del_buyer', methods=['POST','GET'])
    def del_buyer():
        form_data = request.json
        inx = form_data['inx']
        #pp(inx)
        update_db(DB_CONT["BUYER"]["path"], delete_query("BUYER",'pk'), [inx])
        
        return make_response(redirect('viewDB'))
        
                #update_db(db_path, delete_query(table_name, 'OrderNo'), [form_data[0]])
#####


    @app.route('/save_db/', methods=['POST', 'GET'])
    def save_db():

        form_data = request.json

        insert_values = []
        # data matching
        """        
        insert_values[0] = str(form_data[1]) + str(form_data[2])
        insert_values[1] = form_data[86]
        insert_values[2] = form_data[0]
        insert_values[3] = form_data[3]
        insert_values[4] = form_data[39]
        insert_values[5] = form_data[26]
        insert_values[6] = form_data[27]
        insert_values[7] = form_data[28]
        insert_values[8] = form_data[29]
        insert_values[9] = form_data[30]  # remark   연결 안됨
        insert_values[10] = form_data[43][0]  # maker   연결 안됨
        insert_values[11] = form_data[33]  # nego 수량
        insert_values[12] = form_data[38]  # nego 달러
        insert_values[13] = form_data[70]  # nego 원화
        insert_values[14] = form_data[43][3]  #fob amount 달러  연결 안됨
        insert_values[15] = 0  # fob amount 원화  연결안됨
        insert_values[16] = form_data[57][5]  # 선명
        insert_values[17] = form_data[52]  # 서류일자
        insert_values[18] = form_data[53]  # 선적일자
        insert_values[19] = form_data[113]  # 선임달라  연결 안됨   local 선임
        insert_values[20] = form_data[114] # 선임원화
        insert_values[21] = form_data[119]  # 적화보험료달러
        insert_values[22] = form_data[120] # 적화보험료원화
        insert_values[23] = form_data[117]  # 수출보험료달러
        insert_values[24] = form_data[118]  # 수출보험료원화
        insert_values[25] = form_data[131]  # 우편료달러
        insert_values[26] = form_data[132]  # 우편료원화
        insert_values[27] = form_data[123]  # 기타달러
        insert_values[28] = form_data[124]  # 기타원화
        insert_values[29] = form_data[141]  # Margin달러
        insert_values[30] = form_data[142]  # Margin원화
        insert_values[31] = form_data[81]  # 면장번호
        insert_values[32] = form_data[87]  # 면장신고일자
        insert_values[33] = form_data[83]  # 면장FOB 달러
        insert_values[34] = form_data[84]  # 면장FOB 원화
        insert_values[35] = form_data[59]  # Nego 일자
        insert_values[36] = form_data[60]  # Ref No
        insert_values[37] = form_data[66]  # 화폐단위
        insert_values[38] = form_data[67]  # Nego 기타통화
        insert_values[39] = 0  # 인수일자   연결안됨
        insert_values[40] = form_data[88]  # 인수확정일자  연결안됨
        insert_values[41] = 0  # PaymentDate  # 연결안됨
        insert_values[42] = 0  # 면장번호  중복 DB
        insert_values[43] = form_data[9]  # order payment
        insert_values[44] = form_data[57][1]  # loteNo
        insert_values[45] = form_data[57][0]  # conno
        insert_values[46] = form_data[54]  # BL No
        insert_values[47] = form_data[92][6]  # 결제일자  연결안되어 있는것 같다
        insert_values[48] = 0  # 화폐단위  66번과 중복
        insert_values[49] = form_data[36]  # nego 단가
        insert_values[50] = form_data[43][2]  # fob 단가
        insert_values[51] = form_data[10]  # term
        insert_values[52] = 0  # bank  연동 안됨
        insert_values[53] = 0  # invoice no  연동 안됨
        insert_values[54] = 0  # invoice date  연동 안됨
        insert_values[55] = form_data[8]  # 도착지
        insert_values[56] = 0  # 도착지  연동 안됨
        insert_values[57] = 0  # BuyerBankName  연동 안됨
        insert_values[58] = 0  # 잔액외화  연동 안됨
        insert_values[59] = 0  # 잔액달러  연동 안됨
        insert_values[60] = 0  # TT   연동 안됨
        insert_values[61] = form_data[13]  # com1
        insert_values[62] = form_data[14]  # amount1
        insert_values[63] = form_data[15]  # agent
        insert_values[64] = form_data[16]  # com2
        insert_values[65] = form_data[17]  # amount2
        insert_values[66] = form_data[18]  # agent
        insert_values[67] = form_data[19]  # com3
        insert_values[68] = form_data[20]  # amount3
        insert_values[69] = form_data[21]  # agent
        insert_values[70] = form_data[22]  # com4
        insert_values[71] = form_data[23]  # amount4
        insert_values[72] = form_data[24]  # agent
        insert_values[73] = form_data[75]  # 환가요율
        insert_values[74] = form_data[51] # old_orderno
        insert_values[75] = 0  # less deley 달러  연동 안됨
        insert_values[76] = 0  # less deley 원화  연동 안됨
        insert_values[77] = form_data[90]  # less 달러
        insert_values[78] = 0  # less 원화  연동 안됨
        insert_values[79] = 0  # 인수지연달러  연동 안됨
        insert_values[80] = 0  # 인수지연원화  연동 안됨
        insert_values[81] = form_data[93]  # 계산서일자1
        insert_values[82] = form_data[94]  # 계산서수량1
        insert_values[83] = form_data[95]  # 계산서금액1
        insert_values[84] = form_data[96]  # 계산서금액달러1
        insert_values[85] = form_data[97]  # 계산서일자2
        insert_values[86] = form_data[98]  # 계산서수량2
        insert_values[87] = form_data[99]  # 계산서금액2
        insert_values[88] = form_data[100]  # 계산서금액달러2
        insert_values[89] = form_data[101]  # 계산성일자3
        insert_values[90] = form_data[102]  # 계산서수량3
        insert_values[91] = form_data[103]  # 계산서금액3
        insert_values[92] = form_data[104]  # 계산서금액달러3
        insert_values[93] = form_data[105]  # 계산성일자4
        insert_values[94] = form_data[106]  # 계산서수량4
        insert_values[95] = form_data[107]  # 계산서금액4
        insert_values[96] = form_data[108]  # 계산서금액달러4
        insert_values[97] = form_data[57][2]  # 출고일
        insert_values[98] = form_data[89][0]  # 입금일자1
        insert_values[99] = form_data[89][1]  # 입금액1
        insert_values[100] = 0  # 입금일자2
        insert_values[101] = 0  # 입금액2
        insert_values[102] = 0  # 입금일자3
        insert_values[103] = 0  # 입금액3
        insert_values[104] = 0  # 입금일자4
        insert_values[105] = 0  # 입금액4
        insert_values[106] = 0  # 입금일자5
        insert_values[107] = 0  # 입금액5
        insert_values[108] = form_data  # 환가료환출달러  연결 안됨
        insert_values[109] = form_data  # 환가료환출원화  연결 안됨
        insert_values[110] = form_data[32]  # PONUMBER   연결 안됨
        insert_values[111] = form_data[44]  # order 예상선임
        insert_values[112] = form_data[115]  # demurrage 달러
        insert_values[113] = form_data[116]  # demurrage 원화
        insert_values[114] = 0  # undervalue  연결 안됨
        """

        ## DB
        data = DB_CONT
        ##

        length = len(form_data)
        iddb = form_data[length-2]
        idgb = form_data[length-1]
        idstring = str(iddb) + '_' + str(idgb)
        form_data[1] = iddb
        form_data[2] = idgb
        
        del form_data[length-1]
        
        del form_data[length-2]
        

        iCnt=0

        where_var = 'OrderNo="'+iddb+'"'
        where_var2 = 'OrderGubun="' + idgb + '"'
        path = 'dv/data/db/alldata.db'
        table_name = 'ALLDATA'
        select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))
        if len(select_new) != 0:
            msg = '[error] there is same DB name.'
            return jsonify({'msg':msg})
        #if os.path.exists('newdb.list') :
        #    fileread=open('newdb.list','r')
        #    lines = fileread.readlines()
        #    for line in lines:
        #        iCnt+=1         
        #        if idstring in line:
        #            msg = '[error] there is same DB name.'
        #            return jsonify({'msg':msg})


        iCnt = query_db(data['alldata']['path'], select_query(data['alldata']['table_name'], 'count(*)'))[0][0]
        

        if iCnt is None:
            iCnt = 0

        ## write
        #file = open('newdb.list', 'a')
        # file.write(str(form_data[1]+'\n'));
        #file.write(str(form_data[1] + '_' + form_data[2] + '\n'))
        #file.close()

        ## save DB
        iCnt = iCnt+1
        insert_values.append(iCnt)
        i = 0
        for item in form_data:
            i += 1
            tablevalue = ''
            if i == 44 or i == 58 or i == 90 or i == 93:
                for item_sub in item:
               		tablevalue += str(item_sub) + ';'
                item = tablevalue

            insert_values.append(item)


        
        ##### Add new data to a DB

        DB_NM = 'alldata'
        db_path = data[DB_NM]['path']
        table_name = data[DB_NM]['table_name']
        col_num = 149
        column_list = data[DB_NM]['column']
        col_list = ''
        for item in column_list:        	
            if (item == 'pk') | (item=='port'):
                continue
            col_list += '"' + item + '",'

        col_list = col_list[:-1]

        retmsg = update_db(db_path, insert_query(table_name, '(' + col_list + ')', col_num), insert_values)
        if retmsg == 'rollback' :
            return jsonify({'msg': 'error in saving DB'})


        ## save user 
        ## save user
        userpath = data['alldata_user']['path']
        where_user = 'order_title="'+str(form_data[1] + form_data[2])+'"'
        #col_list_user = ['order_title','user_name']
        column_list_user = data['alldata_user']['column']
        col_list_user = ''
        for item2 in column_list_user:            
            col_list_user += '"' + item2 + '",'

        col_list_user = col_list_user[:-1]

        user_value = [str(form_data[1] + form_data[2]), str(session['name'])]
        retmsg = update_db(userpath, insert_query('alldata_user', '(' + col_list_user + ')', 2), user_value)
        if retmsg == 'rollback' :
            return jsonify({'msg': 'error in saving DB'})
        ## end save user

        msg = 'Save complete'
        return jsonify({'msg': msg})
        ######

    @app.route('/delete_db/', methods=['POST', 'GET'])
    def delete_db():

        form_data = request.json

        data = DB_CONT
        DB_NM = 'alldata'
        table_name = data[DB_NM]['table_name']
        db_path = data[DB_NM]['path']

        #update_db(db_path, delete_query(table_name, 'OrderNo'), [form_data[0]])

        where_var = 'OrderNo="' + form_data[0] + '"'
        where_var2 = 'OrderGubun="' + form_data[1] + '"'

        stringDB = str(form_data[0])+'_'+str(form_data[1])

        update_db(db_path, delete_query_two(table_name, where_var, where_var2))

        #if os.path.exists('newdb.list') :
        #    fileread=open('newdb.list','r')

        #    lines = fileread.readlines()
        #f = open('newdb.list','w')
        #for line in lines :
        #    if line.strip("\n") != str(stringDB):
        #        f.write(line)


        return jsonify({'msg': 'Delete complete'})
    
    @app.route('/save_contract_invoice/', methods=['POST', 'GET'])
    def save_contract_invoice():

        
        form_data = request.json
        
        ## DB
        data = DB_CONT
        ##

        iddb = form_data[0]
        idgb = form_data[1]
        portdata = form_data[2]
        #idstring = str(iddb) + '_' + str(idgb)

        iCnt=0
        #if os.path.exists('newdb.list') :
        #    fileread=open('newdb.list','r')

        #    lines = fileread.readlines()
        #    for line in lines:
        #        iCnt+=1                
        #        if idstring in line:
        #            print('found DB:' + str(idstring))


        iCnt = query_db(data['alldata']['path'], select_query(data['alldata']['table_name'], 'count(*)'))[0][0]
        if iCnt is None:
            iCnt = 0

        DB_NM = 'alldata'
        db_path = data[DB_NM]['path']
        table_name = data[DB_NM]['table_name']        

        update_db(db_path, update_query2(table_name, 'port', 'OrderNo', 'OrderGubun'), [portdata,iddb,idgb])


        msg = 'Save complete'
        return jsonify({'msg': msg})

    @app.route('/save_contract2/', methods=['POST', 'GET'])
    def save_contract2():
        form_data = request.json
        
        ## DB
        data = DB_CONT
        ##

        vectordata = [0 for _ in range(31)]
        vectordata[0] = iddb = form_data[0]
        vectordata[1] = idgb = form_data[1]
        vectordata[2] = contract2_invoice_shipper = form_data[2]
        vectordata[3] = date_contract2_ship = form_data[3]
        vectordata[30] = contract2_loading = form_data[4]
        vectordata[5] = contract2_insurance = form_data[5]
        vectordata[7] = date_contract2_contract_date = form_data[6]
        vectordata[8] = contract2_no  = form_data[7]
        vectordata[9] = contract2_reference = form_data[8]
        vectordata[10] = contract2_discharging = form_data[9]
        vectordata[11] = contract2_trans = form_data[10]
        vectordata[12] = contract2_shipment = form_data[11]
        vectordata[13] = contract2_payment = form_data[12]
        vectordata[6] = contract2_remark = form_data[13]
        vectordata[29] = contract2_term = form_data[14]
        vectordata[15] = contract2_goods1 = form_data[15]
        vectordata[19] = contract2_goods2 = form_data[16]
        vectordata[23] = contract2_goods3 = form_data[17]
        vectordata[16] = contract2_quantity_1 = form_data[18]
        vectordata[20] = contract2_quantity_2 = form_data[19]
        vectordata[24] = contract2_quantity_3 = form_data[20]
        vectordata[27] = contract2_quantity_total = form_data[21]
        vectordata[17] = contract2_price_1 = form_data[22]
        vectordata[21] = contract2_price_2 = form_data[23]
        vectordata[24] = contract2_price_3 = form_data[24]
        vectordata[18] = contract2_amount_1 = form_data[25]
        vectordata[22] = contract2_amount_2 = form_data[26]
        vectordata[26] = contract2_amount_3 = form_data[27]
        vectordata[28] = contract2_amount_total = form_data[28]

        
        #idstring = str(iddb) + '_' + str(idgb)


        iCnt = query_db(data['alldata_contract2']['path'], select_query(data['alldata_contract2']['table_name'], 'count(*)'))[0][0]
        if iCnt is None:
            iCnt = 0

        

        DB_NM = 'alldata_contract2'
        db_path = data[DB_NM]['path']
        table_name = data[DB_NM]['table_name']        
        
        col_num = 31
        column_list = data[DB_NM]['column']
        col_list = ''
        for item in column_list:            
            col_list += '"' + item + '",'
        col_list = col_list[:-1]

        where_var = 'ordercon="'+iddb+'"'
        where_var2 = 'gubun="' + idgb + '"'
        select_new = query_db(db_path, selection_query_two(table_name,'*',where_var,where_var2))

        
        retmsg = ''
        if len(select_new) :             
            retmsg=update_db(db_path, update_query2(table_name, 'messrs', 'ordercon', 'gubun'), [contract2_invoice_shipper,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'shipment', 'ordercon', 'gubun'), [date_contract2_ship,iddb,idgb])
            
            retmsg=update_db(db_path, update_query2(table_name, 'insurance', 'ordercon', 'gubun'), [contract2_insurance,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'remark', 'ordercon', 'gubun'), [contract2_remark,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'contractdate', 'ordercon', 'gubun'), [date_contract2_contract_date,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'contractno', 'ordercon', 'gubun'), [contract2_no,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'buyersref', 'ordercon', 'gubun'), [contract2_reference,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'portodfdischarging', 'ordercon', 'gubun'), [contract2_discharging,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'transshipment', 'ordercon', 'gubun'), [contract2_trans,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'partialshipment', 'ordercon', 'gubun'), [contract2_shipment,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'payment"', 'ordercon', 'gubun'), [contract2_payment,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'commodity1', 'ordercon', 'gubun'), [contract2_goods1,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'quantity1', 'ordercon', 'gubun'), [contract2_quantity_1,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'unitprice1', 'ordercon', 'gubun'), [contract2_price_1,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'about1', 'ordercon', 'gubun'), [contract2_amount_1,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'commodity2', 'ordercon', 'gubun'), [contract2_goods2,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'quantity2', 'ordercon', 'gubun'), [contract2_quantity_2,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'unitprice2', 'ordercon', 'gubun'), [contract2_price_2,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'about2', 'ordercon', 'gubun'), [contract2_amount_2,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'commodity3', 'ordercon', 'gubun'), [contract2_goods3,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'quantity3', 'ordercon', 'gubun'), [contract2_quantity_3,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'unitprice3', 'ordercon', 'gubun'), [contract2_price_3,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'about3', 'ordercon', 'gubun'), [contract2_amount_3,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'totalquantity', 'ordercon', 'gubun'), [contract2_quantity_total,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'totalabout', 'ordercon', 'gubun'), [contract2_amount_total,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'term', 'ordercon', 'gubun'), [contract2_term,iddb,idgb])
            retmsg=update_db(db_path, update_query2(table_name, 'port', 'ordercon', 'gubun'), [contract2_loading,iddb,idgb])


        else :
            retmsg=update_db(db_path, insert_query(table_name, '(' + col_list + ')', col_num), vectordata)

        if retmsg == 'rollback' :
            msg = 'failed update db'
        else :
            msg = 'Save complete' 

        
        return jsonify({'msg': msg})

    

    def get_contract_invoice():
    	returndata = []
    	if os.path.exists('save_invoice.dat'):
    		with open('save_invoice.dat','r') as fh:
    			lines = fh.readlines()
    			for line in lines:
    				line = line.replace('\n','')
    				returndata.append(line)
    	return returndata
    		

    @app.route('/modify_db/', methods=['POST', 'GET'])
    def modify_db():
        form_data = request.json

        length = len(form_data)
        iddb = form_data[length-2]
        idgb = form_data[length-1]
        idstring = str(iddb) + '_' + str(idgb)
        form_data[1] = iddb
        form_data[2] = idgb
        
        del form_data[length-1]        
        del form_data[length-2]        
        
        check = False

        where_var = 'OrderNo="'+iddb+'"'
        where_var2 = 'OrderGubun="' + idgb + '"'
        path = 'dv/data/db/alldata.db'
        table_name = 'ALLDATA'
        select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))
        if len(select_new) != 0:
            check = True
        ## check DB ## 
        #if os.path.exists('newdb.list') :
        #    fileread=open('newdb.list','r')
        #    lines = fileread.readlines()
        #    for line in lines:        
        #        if idstring in line:
        #            check = True
                    
        if check == False:
            #return jsonify({'msg' : '[error] can not modify old DB'})
            return jsonify({'msg' : '[Save!] 먼저 상단의 Save를 하십시요! '})
        ##############

        data = DB_CONT
        DB_NM = 'alldata'
        table_name = data[DB_NM]['table_name']
        db_path = data[DB_NM]['path']

        db_rows = data[DB_NM]['column']

        #update_db(USER_LIST, update_query('user_list',['cookie_id'],'num'),[None,num])

        for i  in range(len(db_rows)):
            if i == 0 :
                continue
            #print(range(len(db_rows)))
            #print('item', db_rows[i], 'OrderNo', iddb, 'OrderGubun', idgb, 'data',form_data[i-1] , 'i:', i)            

            tablevalue = ''
            if i == 44 or i == 58 or i == 90 or i == 93:
                for item_sub in form_data[i-1]:
                    tablevalue += str(item_sub) + ';'
                form_data[i-1] = tablevalue

            if i >= 149:
                continue
							
            update_db(db_path, update_query2(table_name, db_rows[i], 'OrderNo', 'OrderGubun'), [form_data[i-1],iddb,idgb])
        



        return jsonify({'msg': 'Modify complete'})

    @app.route('/get_buyer_db/', methods=['POST', 'GET'])
    def get_buyer_db():
        data = DB_CONT
        form_data = request.json # buyer name
        # buyer
        buyerselected = form_data
        path_buyer = data['BUYER']['path']
        table_buyer = 'BUYER'
        where_buyer = '프린트Buyer="'+str(buyerselected)+'"'
        select_buyer = query_db(path_buyer,selection_query(table_buyer,'*',where_buyer))
        # end buyer

        return jsonify({'select_buyer':select_buyer})
    
    @app.route('/get_finaldestination_db/', methods=['POST', 'GET'])
    def get_finaldestination_db():
        data = DB_CONT
        form_data = request.json # buyer name
        # buyer
        buyerselected = form_data
        path_buyer = data['DESTINATION']['path']
        table_buyer = 'DESTINATION'
        where_buyer = '프린트도착지1="'+str(buyerselected)+'"'
        destination = query_db(path_buyer,selection_query(table_buyer,'*',where_buyer))
        # end buyer

        return jsonify({'destination':destination})

    
    @app.route('/viewDB_ret/', methods=['POST','GET'])    
    def viewDB_ret():
        try:
            departdate=request.args.get('departdate')
            returndate=request.args.get('returndate')
            
            total_db = query_db(DB_CONT["alldata"]["path"], select_query("alldata"))
            limit_cnt = 0
            final_data = []
            final_db ={}

            f = request.json

            if f:
                final_data, list_count, view_db, db_name = query_db_all_sel(DB_CONT["alldata"]["view_db"],f)
                return jsonify({'data':final_data,'list_count':list_count, 'view_db':view_db, 'db_name':db_name})
            else:
                limit_cnt = '0,100'

            final_db,db_data,total_count,data_list = init_query_db_all(DB_CONT["alldata"]["view_db"], DB_CONT['alldata']['path'], limit_cnt, 'alldata')


            searchdata = ['buyer_order','OrderNo','OrderGubun','agent_order','item_order','order_DN','order_LF','LS_order','order_remark','order_quantity','QTY_order',
            'order_prc','order_dollor','table_supply']

            total_db_final = []
            # date check

            
            for item in total_db:
                match = re.search(r'\d{4}-\d{2}-\d{2}', item[5])                # 기준키
                if item[5] is not '' : 
                    a = datetime.strptime(item[5], "%Y-%m-%d")
                    b = datetime.strptime(departdate, "%Y-%m-%d")
                    c = datetime.strptime(returndate, "%Y-%m-%d")

                    if (a>b) & (a<c) :

                        newitem =[] 
                        newitem.append(item[0]) # index
                        newitem.append(item[2]) # order no
                        newitem.append(item[3]) # gubun                        
                        newitem.append(item[1]) # buyer
                        newitem.append(item[4]) # agent
                        newitem.append(item[27]) # item 
                        newitem.append(item[28]) # DN 
                        newitem.append(item[29]) # LF 
                        newitem.append(item[30]) # LS 
                        newitem.append(item[31]) # remark 
                        newitem.append(str(item[34]).replace(",","")) # QTY 
                        newitem.append(item[35]) # 수량단위
                        newitem.append(item[37]) # U/PRX
                        newitem.append(str(item[39]).replace(",","")) # AMOUNT
                        #newitem.append(item[]) # FOB단가
                        #newitem.append(item[]) # FOB amount
                        item44=item[44].split(';')                        
                        len1=len(item44)
                        rowlen=len1/4
                        t1 = item44[2]
                        t2 = item44[3]
                        newitem.append(str(t1).replace(",","")) # FOB단가
                        newitem.append(str(t2).replace(",","")) # FOB amount
                        newitem.append(item[7]) # DELI
                        newitem.append(item[8]) # orderterm
                        newitem.append(item[9]) # 도착지
                        newitem.append(item[10]) # orderPayment
                        newitem.append(item[11]) # days

                        i1=0
                        i2=0
                        i3=0
                        i4=0
                        if item[14] == '' :
                            i1 = 0
                        else :
                            i1 = item[14]
                        if item[17] == '' :
                            i2 = 0
                        else  :
                            i2 = item[17]
                        if item[20] == '' :
                            i3 = 0
                        else :
                            i3 = item[20]
                        if item[23] == '' :
                            i4 = 0
                        else:
                            i4 = item[23]

                        t1 =float(i1)+float(i2)
                        t2 =float(i3)+float(i4)
                        newitem.append(str(t1)) # com1
                        newitem.append(str(t2)) # com2                        
                        newitem.append(item[40]) # LCNo
                        newitem.append(item[42]) # SD
                        newitem.append(item[43]) # ED
                        newitem.append(str(item[48]).replace(",","")) # MARGIN
                        newitem.append(item[26]) # OrderCommTotal
                        newitem.append(item[5]) # 계약일자
                        newitem.append(item[36]) # Order화폐단위
                        newitem.append(str(item[39]).replace(",","")) # 외화
                        

                        
                        
                        total_db_final.append(newitem)
                
                
            
            db_data_final =[]
            # data cleanup
            """
            for item in db_data :
                if item in searchdata :
                    if item == 'table_supply' : 
                        db_data_final.append('FOB 단가')
                        db_data_final.append('Fob Amt')
                    elif item == 'buyer_order' :
                        db_data_final.append('Buyer')
                    elif item == 'agent_order' :
                        db_data_final.append('Agent')
                    elif item == 'item_order' :
                        db_data_final.append('Item')
                    elif item == 'order_DN' :
                        db_data_final.append('DN')
                    elif item == 'order_LF' :
                        db_data_final.append('LF')
                    elif item == 'LS_order' :
                        db_data_final.append('LS')
                    elif item == 'order_remark' :
                        db_data_final.append('Remark')
                    elif item == 'order_quantity' :
                        db_data_final.append('QTY')
                    elif item == 'QTY_order' :
                        db_data_final.append('수량단위')
                    elif item == 'order_prc' :
                        db_data_final.append('U/PRX')
                    elif item == 'order_dollor' :
                        db_data_final.append('AMOUNT')
                    elif item == 'date_delivery' : 
                        db_data_final.append('DELI')
                    elif item == 'term_order' :
                        db_data_final.append('OrderTerm')
                    elif item == 'arrive_order' : 
                        db_data_final.append('도착지')
                    elif item == 'payment_order' :
                        db_data_final.append('OrderPayment')
                    elif item == 'order_term' :
                        db_data_final.append('DAYS')
                    elif item == 'order_com1':
                        db_data_final.append('COM1')
                    elif item == 'order_com2':
                        db_data_final.append('COM2')
                    elif item == 'order_LC':
                        db_data_final.append('LCNo')
                    elif item == 'date_SD':
                        db_data_final.append('SD')
                    elif item == 'date_ED':
                        db_data_final.append('ED')
                    elif item == 'order_expect_margin':
                        db_data_final.append('MARGIN')
                    elif item == 'order_com_total':
                        db_data_final.append('OrderCommTotal')
                    elif item == 'datepicker':
                        db_data_final.append('계약일자')
                    elif item == 'PRC_order':
                        db_data_final.append('ORDER화폐단위')
            """  
            db_data_final.append('OrderNo')
            db_data_final.append('OrderGubun')            
            db_data_final.append('Buyer')
            db_data_final.append('Agent')
            db_data_final.append('Item')
            db_data_final.append('DN')
            db_data_final.append('LF')                    
            db_data_final.append('LS')
            db_data_final.append('Remark')
            db_data_final.append('QTY')
            db_data_final.append('수량단위')
            db_data_final.append('U/PRX')
            db_data_final.append('AMOUNT')
            db_data_final.append('FOB 단가')
            db_data_final.append('Fob Amt')
            db_data_final.append('DELI')
            db_data_final.append('OrderTerm')
            db_data_final.append('도착지')
            db_data_final.append('OrderPayment')
            db_data_final.append('DAYS')
            db_data_final.append('COM1')
            db_data_final.append('COM2')
            db_data_final.append('LCNo')
            db_data_final.append('SD')
            db_data_final.append('ED')
            db_data_final.append('MARGIN')
            db_data_final.append('OrderCommTotal')
            db_data_final.append('계약일자')
            db_data_final.append('ORDER화폐단위')
            db_data_final.append('외화')


            
                
            return render_template('viewdb_ret.html', \
                                    total_db=total_db_final, db_data=db_data_final)
        except TemplateNotFound:
            abort(404)

    @app.route('/search/', methods=['POST','GET'])
    def search():
        try:

            ###########################################################################
            ## chrolling
            ###########################################################################
            '''
            driver = wd.Chrome('C:\\Users\\sweet\\Downloads\\chromedriver')            
            driver.get("https://biz.kebhana.com/foex/rate/index.do?menuItemId=wcfxd740_101i%22")
            time.sleep(1)
        
            elem = driver.find_element_by_xpath("//input[@id='tmpDt']")
            elem.clear()
            elem.send_keys('2019-05-27' + Keys.RETURN)
            driver.find_element_by_xpath('//*[@id="hanaMainForm"]/div/table/tbody/tr[2]/td/label[3]/button').click()
            driver.find_element_by_xpath('//*[@id="hanaMainDiv"]/fieldset/div[2]/a').click()
            
            time.sleep(1)
            
            html = driver.execute_script('return document.body.innerHTML')
            soup = BeautifulSoup(html, "html.parser")
            
            cols = soup.findAll('td')            
            for idx, col in enumerate(cols):
                if '미국'in col.text :
                    textfinal=cols[idx+1]                    
                    textfinal  = str(textfinal)
                    text = re.sub('<.+?>', '', textfinal, 0, re.I|re.S)
                    print (text)
                    
            driver.quit()
            '''
            ###########################################################################

            ###########################################################################
            ## search exchange 
            ###########################################################################
            res=requests.post('https://www.kebhana.com/cms/rate/wpfxd651_01i_01.do', data={'curCd':'USD', 'pbldDvCd':1, 'inqStrDt':20190527})            
            soup=BeautifulSoup(res.text,"html.parser")
            cols = soup.findAll('td')
            for idx, col in enumerate(cols):                
                if '미국'in col.text :
                    textfinal=cols[idx+9]                    
                    textfinal  = str(textfinal)
                    text = re.sub('<.+?>', '', textfinal, 0, re.I|re.S)
                    
            ###########################################################################


            total_db = query_db(DB_CONT["alldata"]["path"], select_query("alldata"))
            limit_cnt = 0
            final_data = []
            final_db ={}

            f = request.json

            if f:
                final_data, list_count, view_db, db_name = query_db_all_sel(DB_CONT["alldata"]["view_db"],f)
                return jsonify({'data':final_data,'list_count':list_count, 'view_db':view_db, 'db_name':db_name})
            else:
                limit_cnt = '0,100'

            final_db,db_data,total_count,data_list = init_query_db_all(DB_CONT["alldata"]["view_db"], DB_CONT['alldata']['path'], limit_cnt, 'alldata')

            searchdata = ['buyer_order','OrderNo','OrderGubun','agent_order','item_order','order_DN','order_LF','LS_order','order_remark','order_quantity','QTY_order',
            'order_prc','order_dollor','table_supply']
            
            db_data_final =[]
            # data cleanup
            for item in db_data : 
                if item in searchdata :
                    if item == 'table_supply' : 
                        db_data_final.append('FOB 단가')
                        db_data_final.append('Fob Amt')
                    elif item == 'buyer_order' :
                        db_data_final.append('Buyer')
                    elif item == 'agent_order' :
                        db_data_final.append('Agent')
                    elif item == 'item_order' :
                        db_data_final.append('Item')
                    elif item == 'order_DN' :
                        db_data_final.append('DN')
                    elif item == 'order_LF' :
                        db_data_final.append('LF')
                    elif item == 'LS_order' :
                        db_data_final.append('LS')
                    elif item == 'order_remark' :
                        db_data_final.append('Remark')
                    elif item is 'order_quantity' :
                        db_data_final.append('QTY')
                    elif item is 'QTY_order' :
                        db_data_final.append('수량단위')
                    elif item is 'order_prc' :
                        db_data_final.append('U/PRX')
                    elif item is 'order_dollor' :
                        db_data_final.append('AMOUNT')
                    else : 
                        db_data_final.append(item)

                
            
            return render_template('search.html', \
                                    total_db=total_db, db_data=db_data_final)
        except TemplateNotFound:
            abort(404)

    @app.route('/search2/', methods=['POST','GET'])
    def search2():
        try:
    
            return render_template('search2.html')
        except TemplateNotFound:
            abort(404)

    @app.route('/viewDB_ret_search2/', methods=['POST','GET'])    
    def viewDB_ret_search2():
        try:

            #print('start')
            departdate=request.args.get('checkin')
            returndate=request.args.get('checkout')

            #print('departdate')
            #print(departdate)
            #print(returndate)
            
            total_db = query_db(DB_CONT["alldata"]["path"], select_query("alldata"))

            
            limit_cnt = 0
            final_data = []
            final_db ={}

            #print(request.json)
            #f_ret2 = request.json
            #print(request.json)
            

            if request.get_data():
                final_data, list_count, view_db, db_name = query_db_all_sel(DB_CONT["alldata"]["view_db"],f)
                return jsonify({'data':final_data,'list_count':list_count, 'view_db':view_db, 'db_name':db_name})
            else:
                limit_cnt = '0,100'

            print('final_data')

            final_db,db_data,total_count,data_list = init_query_db_all(DB_CONT["alldata"]["view_db"], DB_CONT['alldata']['path'], limit_cnt, 'alldata')


            total_db_final = []
            # date check
            for item in total_db:
                match = re.search(r'\d{4}-\d{2}-\d{2}', item[81])                # nego 예정일 
                if item[81] is not '' : 
                    a = datetime.strptime(item[81], "%Y-%m-%d")
                    b = datetime.strptime(departdate, "%Y-%m-%d")
                    c = datetime.strptime(returndate, "%Y-%m-%d")
                    
                    
                    if (a>b) & (a<c) :
                        newitem =[] 
                        newitem.append(item[0]) # index       
                        newitem.append(item[2]) # OrderNo     0
                        newitem.append(item[87]) #              1
                        newitem.append(item[1]) # buyer   2
                        newitem.append(item[4]) # agent 3
                        newitem.append(item[40]) # LCNo 4
                        newitem.append(item[27]) # Item 5
                        newitem.append(item[28]) # DN 6
                        newitem.append(item[29]) # LF 7
                        newitem.append(item[30]) # LS 8 
                        newitem.append(item[31]) # Remark  9
                        
                        item44=item[44].split(';')
                        len1=len(item44)
                        rowlen=len1/5
                        t1 = item44[0]                        

                        newitem.append(t1) # Maker   10

                        
                        newitem.append(str(item[63]).replace(",","")) # Nego 수량  11
                        newitem.append(str(item[73]).replace(",","")) # Nego달러 12                        
                        newitem.append(str(item[71]).replace(",","")) # Nego 우ㅓㄴ호ㅏ  13

                        item93 = item[93].split(';')
                        pp(item93)
                        len1 =len(item93)
                        
                        rowlen = (len1-1)/9
                        
                        sumitem1=0 
                        sumitem2=0
                

                        for k in range(int(rowlen)):
                            sumitem1 = item93[4+k*8]
                            sumitem2 = item93[4+k*9]


                        newitem.append(str(sumitem1).replace(",","")) # FOBAmount달러 14
                        newitem.append(str(sumitem2).replace(",","")) # FOBAmount원화  15
                        
                        item58 = item[58].split(';')
                        rowlen=len(item58)
                        t1 = item58[5]
                        
                        newitem.append(t1) # 선명  16
                        newitem.append(item[53]) # 서류일자 17
                        newitem.append(item[54]) # 선적일자  18
                        newitem.append(str(item[114]).replace(",","")) # 선임달러 19
                        newitem.append(str(item[115]).replace(",","")) # 선임원화  20
                        newitem.append(str(item[120]).replace(",","")) # 적화보험료달러  21
                        newitem.append(str(item[121]).replace(",","")) # 적화보험료원화  22
                        newitem.append(str(item[118]).replace(",","")) # 수출보험료달러  23
                        newitem.append(str(item[119]).replace(",","")) # 수출보험료원화  24
                        newitem.append(str(item[132]).replace(",","")) # 우편료달러   25
                        newitem.append(str(item[133]).replace(",","")) # 우편료원화  26
                        newitem.append(str(item[124]).replace(",","")) # 기타달러  27
                        newitem.append(str(item[125]).replace(",","")) # 기타원화  28
                        newitem.append(str(item[142]).replace(",","")) # Margin달러  29
                        newitem.append(str(item[143]).replace(",","")) # Margin원화  30
                        newitem.append(item[82]) # 면장번호   31
                        newitem.append(item[88]) # 면장신고일자  32
                        newitem.append(str(item[84]).replace(",","")) # 면장FOB달러  33
                        newitem.append(str(item[86]).replace(",","")) # 면장FOB원화  34
                        newitem.append(item[60]) # Nego일자  35
                        newitem.append(item[61]) # Ref.No   36
                        newitem.append(str(item[67])) # 화폐   37
                        newitem.append(str(item[68]).replace(",","")) # Nego기타통화  38
                        newitem.append('') # 인수일자  39
                        newitem.append(item[89]) # 인수확정일자  40
                        newitem.append('') # PAYMENT DATE   41
                        newitem.append('') # OrderPayment   42
                        
                        item58 = item[58].split(';')
                        rowlen=len(item58)
                        t1 = item58[1]
                        t0 = item58[0]
                        
                        newitem.append(t1) # LoteNo    43
                        newitem.append(t0) # ConNo    44
                        newitem.append(item[55]) # BLNo  45
                        
                        item93 = item[93].split(';')
                        t0 = item93[6]
                        
                        newitem.append(t0) # 결제일자  46
                        newitem.append(item[37]) # Nego 단가  47

                        t0=item44[2]

                        newitem.append(str(t0).replace(",","")) # FOB 단가  48
                        newitem.append(item[8]) # Term  49
                        newitem.append(item[59]) # Bank  50
                        newitem.append(item[2]) # InvoiceNo   51
                        newitem.append(item[53]) # InvoiceDate  52
                        newitem.append(item[9]) # 도착지  53
                        newitem.append(str(item[68]).replace(",","")) # 잔액외화  54
                        newitem.append(str(item[69]).replace(",","")) # 잔액달러  55
                        newitem.append(str(item[14]).replace(",","")) # Com1  56
                        newitem.append(str(item[15]).replace(",","")) # Amount1  57
                        newitem.append(item[16]) # Agent1  58
                        newitem.append(str(item[17]).replace(",","")) # Com2  59
                        newitem.append(str(item[18]).replace(",","")) # Amount2  60
                        newitem.append(item[19]) # Agent2   61
                        newitem.append(str(item[20]).replace(",","")) # Com3  62
                        newitem.append(str(item[21]).replace(",","")) # Amount3   63
                        newitem.append(item[22]) # Agent3  64
                        newitem.append(str(item[23]).replace(",","")) # Amount4   65
                        newitem.append(str(item[23]).replace(",","")) # Agent4  66
                        newitem.append(item[25]) # Agent4  67
                        
                        newitem.append(item[76]) # 환가요율  68
                        newitem.append(item[52]) # old_orderno   69
                        newitem.append(str(item[136]).replace(",","")) # LessDelay달러  70
                        newitem.append(str(item[137]).replace(",","")) # LessDelay원화  71
                        newitem.append(str(item[138]).replace(",","")) # Less달러    72
                        newitem.append(str(item[139]).replace(",","")) # Less원화    73
                        newitem.append(str(item[140]).replace(",","")) # 인수지연달러   74
                        newitem.append(str(item[141]).replace(",","")) # 인수지연원화   75
                        newitem.append(item[94]) # 계산서일자1    76
                        newitem.append(str(item[98]).replace(",","")) # 계산서수량1    77
                        newitem.append(str(item[102]).replace(",","")) # 계산서금액1    78
                        newitem.append(str(item[106]).replace(",","")) # 계산서금액달러1  79
                        newitem.append(item[95]) # 계산서일자2   80
                        newitem.append(str(item[99]).replace(",","")) # 계산서수량2    81
                        newitem.append(str(item[103]).replace(",","")) # 계산서금액2   82
                        newitem.append(str(item[107]).replace(",","")) # 계산서금액달러2   83
                        newitem.append(item[96]) # 계산서일자3  84
                        newitem.append(str(item[100]).replace(",","")) # 계산서수량3  85
                        newitem.append(str(item[104]).replace(",","")) # 계산서금액3  86
                        newitem.append(str(item[108]).replace(",","")) # 계산서금액달러3  87
                        newitem.append(item[97]) # 계산서일자4  88
                        newitem.append(str(item[101]).replace(",","")) # 계산서수량4   89
                        newitem.append(str(item[105]).replace(",","")) # 계산서금액4    90
                        newitem.append(str(item[109]).replace(",","")) # 계산서금액달러4  91
                        
                        t1 = item58[2]

                        newitem.append(t1) # 출고일   92

                        item90 = item[90].split(';')
                        rowlen=(len(item90)-1)/3

                        for k in range(4):
                            if rowlen <= k:
                                t11=0
                                t12=0
                            else :                                
                                t11 = item90[int(0+k*3)]
                                t12 = item90[int(1+k*3)]                                
                            
                            newitem.append(str(t11)) # 입금일자1  93  95  97  99
                            newitem.append(str(t12).replace(",","")) # 입금액1    94  96  98  100                            
                        

                        newitem.append(str(item[134]).replace(",","")) #환가료환출달러')    #101
                        newitem.append(str(item[135]).replace(",","")) #환가료환출원화')    #102
                        newitem.append(item[33]) #PONUMBER')         #103
                        newitem.append(str(item[45]).replace(",","")) #Order예상선임')     #104
                        newitem.append(str(item[116]).replace(",","")) #Demurrage달러')     #105
                        newitem.append(str(item[117]).replace(",","")) #Demurrage원화')     106
                        newitem.append(str(item[66]).replace(",","")) #UnderValue')      107
                        
                        total_db_final.append(newitem)
            
            db_data_final =[]
            db_data_final.append('InvNo')  #0
            db_data_final.append('DueDate')  #1          
            db_data_final.append('Buyer')  #2
            db_data_final.append('Agent') #3
            db_data_final.append('LCNo') #4
            db_data_final.append('Item') #5
            db_data_final.append('DN')   #6               
            db_data_final.append('LF')   #7
            db_data_final.append('LS')   #8
            db_data_final.append('Remark') #9
            db_data_final.append('Maker')  #10
            db_data_final.append('Nego수량') #11
            db_data_final.append('Nego달러') #12
            db_data_final.append('Nego원화') #13
            db_data_final.append('FOBAmount달러') #14
            db_data_final.append('FOBAmount원화') #15
            db_data_final.append('선명')          #16
            db_data_final.append('서류일자')      #17
            db_data_final.append('선적일자')      #18
            db_data_final.append('선임달러')      #19
            db_data_final.append('선임원화')      #20
            db_data_final.append('적화보험료달러') #21
            db_data_final.append('적화보험료원화') #22
            db_data_final.append('수출보험료달러') #23
            db_data_final.append('수출보험료원화') #24
            db_data_final.append('우편료달러')     #25
            db_data_final.append('우편료원화')     #26
            db_data_final.append('기타달러')       #27
            db_data_final.append('기타원화')       #28
            db_data_final.append('Margin달러')     #29
            db_data_final.append('Margin원화')     #30
            db_data_final.append('면장번호')       #31
            db_data_final.append('면장신고일자')   #32
            db_data_final.append('면장FOB달러')    #33
            db_data_final.append('면장FOB원화')    #34 _ add
            db_data_final.append('Nego일자')       #35
            db_data_final.append('Ref.No')         #36
            db_data_final.append('화폐')           #37
            db_data_final.append('Nego기타통화')    #38
            db_data_final.append('인수일자')        #39
            db_data_final.append('인수확정일자')    #40
            db_data_final.append('PAYMENT DATE')   #41
            db_data_final.append('OrderPayment')   #42
            db_data_final.append('SealNo')         #43
            db_data_final.append('ConNo')          #44
            db_data_final.append('BLNo')           #45
            db_data_final.append('결제일자')        #46
            db_data_final.append('Nego 단가')      #47
            db_data_final.append('FOB 단가')       #48
            db_data_final.append('Term')           #49
            db_data_final.append('Bank')           #50
            db_data_final.append('InvoiceNo')      #51
            db_data_final.append('InvoiceDate')    #52
            db_data_final.append('도착지')          #53
            db_data_final.append('잔액외화')        #54
            db_data_final.append('잔액달러')        #55
            db_data_final.append('Com1')            #56
            db_data_final.append('Amount1')         #57
            db_data_final.append('Agent1')          #58
            db_data_final.append('Com2')            #59
            db_data_final.append('Amount2')         #60
            db_data_final.append('Agent2')          #61
            db_data_final.append('Com3')            #62
            db_data_final.append('Amount3')         #63
            db_data_final.append('Agent3')          #64            
            db_data_final.append('Com4')            #65
            db_data_final.append('Amount4')         #66
            db_data_final.append('Agent4')          #67
            db_data_final.append('환가요율')         #68
            db_data_final.append('old_orderno')      #69
            db_data_final.append('LessDelay달러')    #70
            db_data_final.append('LessDelay원화')    #71
            db_data_final.append('Less달러')         #72
            db_data_final.append('Less원화')         #73
            db_data_final.append('인수지연달러')      #74
            db_data_final.append('인수지연원화')      #75
            db_data_final.append('계산서일자1')       #76
            db_data_final.append('계산서수량1')       #77
            db_data_final.append('계산서금액1')       #78
            db_data_final.append('계산서금액달러1')    #79
            db_data_final.append('계산서일자2')       #80
            db_data_final.append('계산서수량2')       #81
            db_data_final.append('계산서금액2')       #82
            db_data_final.append('계산서금액달러2')    #83
            db_data_final.append('계산서일자3')       #84
            db_data_final.append('계산서수량3')       #85
            db_data_final.append('계산서금액3')       #86
            db_data_final.append('계산서금액달러3')    #87
            db_data_final.append('계산서일자4')       #88
            db_data_final.append('계산서수량4')       #88
            db_data_final.append('계산서금액4')       #90
            db_data_final.append('계산서금액달러4')    #91
            db_data_final.append('출고일')            #92
            db_data_final.append('입금일자1')         #93
            db_data_final.append('입금액1')           #94
            db_data_final.append('입금일자2')         #95
            db_data_final.append('입금액2')           #96
            db_data_final.append('입금일자3')         #97
            db_data_final.append('입금액3')           #98
            db_data_final.append('입금일자4')        #99
            db_data_final.append('입금액4')          #100
            db_data_final.append('환가료환출달러')    #101
            db_data_final.append('환가료환출원화')    #102
            db_data_final.append('PONUMBER')         #103
            db_data_final.append('Order예상선임')     #104
            db_data_final.append('Demurrage달러')     #105
            db_data_final.append('Demurrage원화')     #106
            db_data_final.append('UnderValue')       #107

            #print( 'total_db_final',len(total_db_final[0]) )
            #print( 'db_data_final',len(db_data_final) )
                
            return render_template('viewdb_ret_search2.html', \
                                    total_db=total_db_final, db_data=db_data_final)
        except TemplateNotFound:
            abort(404)

    @app.route('/search3/', methods=['POST','GET'])
    def search3():
        try:
    
            return render_template('search3.html')
        except TemplateNotFound:
            abort(404)

    @app.route('/viewDB_ret_search3/', methods=['POST','GET'])    
    def viewDB_ret_search3():
        try:
            departdate=request.args.get('departdate')
            returndate=request.args.get('returndate')
            
            total_db = query_db(DB_CONT["alldata"]["path"], select_query("alldata"))
            limit_cnt = 0
            final_data = []
            final_db ={}

            #f = request.json




            #if f:
            if request.get_data():
                final_data, list_count, view_db, db_name = query_db_all_sel(DB_CONT["alldata"]["view_db"],f)
                return jsonify({'data':final_data,'list_count':list_count, 'view_db':view_db, 'db_name':db_name})
            else:
                limit_cnt = '0,100'

            final_db,db_data,total_count,data_list = init_query_db_all(DB_CONT["alldata"]["view_db"], DB_CONT['alldata']['path'], limit_cnt, 'alldata')


            total_db_final = []
            # date check
            for item in total_db:

                #print(item[58])

                item58 = item[58].split(';')
                rowlen=len(item58)
                t2 = item58[3]                        

                #print('t2',t2)

                match = re.search(r'\d{4}-\d{2}-\d{2}', t2)                # nego 예정일 

                #print('match',match)
                if t2 is not '' : 
                    a = datetime.strptime(t2, "%Y-%m-%d")
                    b = datetime.strptime(departdate, "%Y-%m-%d")
                    c = datetime.strptime(returndate, "%Y-%m-%d")
                    
                    
                    if (a>b) & (a<c) :
                        newitem =[] 
                        newitem.append(item[0]) # index       
                        newitem.append(item[2]) # OrderNo     0
                        newitem.append(item[3]) # OrderGubun  1
                        newitem.append(item[1]) # buyer   2
                        newitem.append(item[4]) # agent 3
                        newitem.append(item[27]) # Item 4
                        newitem.append(item[28]) # DN 5
                        newitem.append(item[29]) # LF 6

                        item44=item[44].split(';')                        
                        t1 = item44[0]
                        newitem.append(t1) # 업체   7

                        item93=item[93].split(';')
                        len1 = len(item93)-1
                        rowlen=len1/9
                        t2 = item93[2]
                        t3 = item93[3]
                        t5 = item93[5]
                        t4 = item93[4]
                        t1 = item93[1]
                        t6 = item93[6]

                        newitem.append(t2) # Local 수량1
                        newitem.append(t3) # LS 8 
                        newitem.append(t5) # Remark  9
                        newitem.append(t4) # Remark  9
                        newitem.append(t1) # Remark  9
                        newitem.append(t6) # Remark  9

                        newitem.append(item[94]) # 계산서일자1
                        newitem.append(item[98]) # 계산서수량1
                        newitem.append(item[106]) # 계산서금액1

                        if rowlen <2 :
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                        else:
                            t2 = item93[2+9*1]
                            t3 = item93[3+9*1]
                            t5 = item93[5+9*1]
                            t4 = item93[4+9*1]
                            t1 = item93[1+9*1]
                            t6 = item93[6+9*1]

                            newitem.append(t2) # Local 수량1
                            newitem.append(t3) # LS 8 
                            newitem.append(t5) # Remark  9
                            newitem.append(t4) # Remark  9
                            newitem.append(t1) # Remark  9
                            newitem.append(t6) # Remark  9
                        
                        newitem.append(item[95]) # 계산서일자2
                        newitem.append(item[99]) # 계산서수량2
                        newitem.append(item[107]) # 계산서금액2

                        if rowlen <3 :
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                        else:
                            t2 = item93[2+9*2]
                            t3 = item93[3+9*2]
                            t5 = item93[5+9*2]
                            t4 = item93[4+9*2]
                            t1 = item93[1+9*2]
                            t6 = item93[6+9*2]

                            newitem.append(t2) # Local 수량3
                            newitem.append(t3) # LS 8 
                            newitem.append(t5) # Remark  9
                            newitem.append(t4) # Remark  9
                            newitem.append(t1) # Remark  9
                            newitem.append(t6) # Remark  9
                        
                        newitem.append(item[96]) # 계산서일자1
                        newitem.append(item[100]) # 계산서수량1
                        newitem.append(item[108]) # 계산서금액1

                        if rowlen <4 :
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                            newitem.append('') 
                        else:
                            t2 = item93[2+9*3]
                            t3 = item93[3+9*3]
                            t5 = item93[5+9*3]
                            t4 = item93[4+9*3]
                            t1 = item93[1+9*3]
                            t6 = item93[6+9*3]

                            newitem.append(t2) # Local 수량3
                            newitem.append(t3) # LS 8 
                            newitem.append(t5) # Remark  9
                            newitem.append(t4) # Remark  9
                            newitem.append(t1) # Remark  9
                            newitem.append(t6) # Remark  9
                        
                        newitem.append(item[97]) # 계산서일자1
                        newitem.append(item[101]) # 계산서수량1
                        newitem.append(item[109]) # 계산서금액1

                        newitem.append(item[60]) # nego 일자
                        newitem.append(item[63]) # nego 수량

                        newitem.append(item[1]) # buyer   2
                        newitem.append(item[4]) # agent 3
                        newitem.append(item[27]) # Item 4

                        newitem.append(item[54]) # 선적일자
                        newitem.append('') # 년월

                        item58 = item[58].split(';')
                        rowlen=len(item58)
                        t2 = item58[2]
                        newitem.append(t2) # 출고일
                        
                        total_db_final.append(newitem)
            
            db_data_final =[]
            db_data_final.append('OrderNo')  #0
            db_data_final.append('OrderGubun')  #1          
            db_data_final.append('Buyer')  #2
            db_data_final.append('Agent') #3

            db_data_final.append('Item') #4
            db_data_final.append('DN')   #5               
            db_data_final.append('LF')   #6
            db_data_final.append('업체')   #7  *
            db_data_final.append('Local 수량1') #8  *
            db_data_final.append('Local 단가1') #9
            db_data_final.append('Local 원화1') #10
            db_data_final.append('Local 달러1') #11
            db_data_final.append('Local LC No1') #12
            db_data_final.append('Local 결제일자1') #13
            db_data_final.append('계산서일자1') #14
            db_data_final.append('계산서수량1') #15
            db_data_final.append('계산서금액1') #16
            db_data_final.append('Local 수량2') #17  *
            db_data_final.append('Local 단가2') #18
            db_data_final.append('Local 원화2') #19
            db_data_final.append('Local 달러2') #20
            db_data_final.append('Local LC No2') #21
            db_data_final.append('Local 결제일자2') #22
            db_data_final.append('계산서일자2') #23
            db_data_final.append('계산서수량2') #24
            db_data_final.append('계산서금액2') #25
            db_data_final.append('Local 수량3') #26
            db_data_final.append('Local 단가3') #27
            db_data_final.append('Local 원화3') #28
            db_data_final.append('Local 달러3') #29
            db_data_final.append('Local LC No3') #30
            db_data_final.append('Local 결제일자3') #31
            db_data_final.append('계산서일자3') #32
            db_data_final.append('계산서수량3') #33
            db_data_final.append('계산서금액3') #34
            db_data_final.append('Local 수량4') #35
            db_data_final.append('Local 단가4') #36
            db_data_final.append('Local 원화4') #37
            db_data_final.append('Local 달러4') #38
            db_data_final.append('Local LC No4') #39
            db_data_final.append('Local 결제일자4') #40
            db_data_final.append('계산서일자4') #41
            db_data_final.append('계산서수량4') #42
            db_data_final.append('계산서금액4') #43

            db_data_final.append('Nego일자') #44
            db_data_final.append('Nego수량') #45
            db_data_final.append('Buyer') #46
            db_data_final.append('Agent') #47
            db_data_final.append('Item') #48
            db_data_final.append('선적일자') #49
            db_data_final.append('년월') #50
            db_data_final.append('출고일') #51
            

            #print( 'total_db_final',len(total_db_final[0]) )
            #print( 'db_data_final',len(db_data_final) )
                
            return render_template('viewdb_ret_search3.html', \
                                    total_db=total_db_final, db_data=db_data_final)
        except TemplateNotFound:
            abort(404)

    
    @app.route('/get_exchange/', methods=['POST','GET'])
    def get_exchange():
      
        form_data = request.json

        datedata=form_data.replace("-","")
        
        valUS=0
        valEU=0

        text=''
        res=requests.post('https://www.kebhana.com/cms/rate/wpfxd651_01i_01.do', data={'curCd':'USD', 'pbldDvCd':1, 'inqStrDt':datedata})
        soup=BeautifulSoup(res.text,"html.parser")
        cols = soup.findAll('td')
        for idx, col in enumerate(cols):                
            if '미국'in col.text :
                textfinal=cols[idx+9]                 
                textfinal  = str(textfinal)
                text = re.sub('<.+?>', '', textfinal, 0, re.I|re.S)                
                valUS = float(text.replace(',',''))
                

        textEUfinal=''
        res=requests.post('https://www.kebhana.com/cms/rate/wpfxd651_01i_01.do', data={'curCd':'EUR', 'pbldDvCd':1, 'inqStrDt':datedata})
        soupEU=BeautifulSoup(res.text,"html.parser")
        cols = soupEU.findAll('td')
        for idx, col in enumerate(cols):                
            if '유로'in col.text :
                textEU=cols[idx+9]                
                textEU  = str(textEU)
                textEUfinal = re.sub('<.+?>', '', textEU, 0, re.I|re.S)
                
                valEU = float(textEUfinal.replace(',',''))

                
                textEUfinal=str(float(valEU) / float(valUS))

        return jsonify({'return':text, 'ret2':textEUfinal})

    @app.route('/save_shipping_db/', methods=['POST', 'GET'])
    def save_shipping_db():

        form_data = request.json
        ## DB
        data = DB_CONT
        ##  

        insert_values=[]
        ## save DB
        i =0 
        for item in form_data:
            if i is 0:
                item = item + form_data[1]
            insert_values.append(item)
            i += 1


        ##### Add new data to a DB

        DB_NM = 'alldata_ship'
        db_path = data[DB_NM]['path']
        table_name = data[DB_NM]['table_name']
        col_num = 34
        column_list = data[DB_NM]['column']
        col_list = ''
        for item in column_list:            
            col_list += '"' + item + '",'

        col_list = col_list[:-1]

        ## check same db
        where = 'order_title="'+str(insert_values[0])+'"'
        temp = query_db(db_path, selection_query(table_name, 'order_title' ,where))
        
        if len(temp) : 
            for j in range(col_num) :
                if (j == 0) | (j == 1):
                    continue
                update_db(db_path, update_query2(table_name, column_list[j],'order_title', 'gubun'), [insert_values[j],insert_values[0],insert_values[1]])            
        else :
            update_db(db_path, insert_query(table_name, '(' + col_list + ')', col_num), insert_values)


        
        
        msg = 'Save complete'
        return jsonify({'msg': msg})

    @app.route('/get_shipping_db/', methods=['POST', 'GET'])
    def get_shipping_db():
        form_data = request.json
        ## DB
        data = DB_CONT
        ##  
        # alldata_ship        
        selected = form_data[0]

        path1 = data['alldata_ship']['path']
        table = 'alldata_ship'
        where = 'order_title="'+str(selected)+'"'
        selecteddata = query_db(path1,selection_query(table,'*',where))        
        # end alldata_ship
        #pp(selecteddata)
        if len(selecteddata):
            msg = 'Get data complete'
            return jsonify({'msg': msg, 'data':selecteddata})
        else :
            return jsonify({'msg': 'Fail to read data'})


    @app.route('/search_jugan/', methods=['POST','GET'])
    def search_jugan():
        try:
            return render_template('search_jugan.html')
        except TemplateNotFound:
            abort(404)

    @app.route('/viewDB_ret_search_jugan/', methods=['POST','GET'])    
    def viewDB_ret_search_jugan():
        try:
            departdate=request.args.get('departdate')
            returndate=request.args.get('returndate')
            
            total_db = query_db(DB_CONT["alldata"]["path"], select_query("alldata"))
            limit_cnt = 0
            final_data = []
            final_db ={}

            #f = request.json


            

        

            if request.get_data():
                final_data, list_count, view_db, db_name = query_db_all_sel(DB_CONT["alldata"]["view_db"],f)
                return jsonify({'data':final_data,'list_count':list_count, 'view_db':view_db, 'db_name':db_name})
            else:
                limit_cnt = '0,100'

            final_db,db_data,total_count,data_list = init_query_db_all(DB_CONT["alldata"]["view_db"], DB_CONT['alldata']['path'], limit_cnt, 'alldata')

            currencydata=[]
            bankdata=[]

            total_db_final = []
            # date check
            
            for item in total_db:

                if session['company'] is 'DB':
                    if 'DB' not in item[2]:
                        continue
                else:
                    if 'EI' not in item[2]:
                        continue



                match = re.search(r'\d{4}-\d{2}-\d{2}', item[81])                # nego 예정일 
                if item[81] is not '' : 
                    a = datetime.strptime(item[81], "%Y-%m-%d")
                    b = datetime.strptime(departdate, "%Y-%m-%d")
                    

                    c = datetime.strptime(returndate, "%Y-%m-%d")
                    
                    
                    if (a>b) & (a<c) :

                        ## DB ## alldata_ship
                        data = DB_CONT                 
                        selected = str(item[2]+item[3])
                        path1 = data['alldata_ship']['path']
                        table = 'alldata_ship'
                        where = 'order_title="'+str(selected)+'"'
                        selecteddata = query_db(path1,selection_query(table,'*',where))

                        

                        username = ''
                        if len(selecteddata) :
                            username = selecteddata[0][29]
                        else :
                            username = session['name']
                        ##

                        newitem =[] 
                        newitem.append(item[0]) # index       
                        newitem.append(item[81]) # 1
                        newitem.append(username) # 2
                        newitem.append(str(item[2] + item[3])) # 3
                        newitem.append(item[1]) # 4
                        newitem.append(item[27]) # 5
                        newitem.append(item[63]) # 6
                        newitem.append(item[72]) # 7
                        newitem.append(item[67]) # 8
                        newitem.append(item[68]) # 9

                        
                        # currency sum
                        if len(currencydata) ==0 :
                            currencydata.append([item[67],item[68]])
                        else :
                            found = False
                            for itemcur in currencydata :
                                if itemcur[0] == item[67] :
                                    found = True

                                    if item[68] == '':
                                        itemcur[1] = float(itemcur[1])
                                    else :
                                        itemcur[1] = float(itemcur[1])+float(item[68])
                            if found == False : 
                                currencydata.append([item[67],item[68]])
                        

                        # payment+days
                        termdata=item[10]+' '+item[11]+'D'
                        newitem.append(termdata) # 10
                        newitem.append(item[59]) # 11

                        # bank data
                        if len(bankdata) == 0:
                            bankdata.append([item[59],[item[67],item[68]]])
                        else :
                            found = False
                            for itemcur in bankdata :                                
                                if itemcur[0] == item[59] : # bank                                    
                                    if len(itemcur[1]) ==0 :
                                        itemcur[1]= [item[67],item[68]]
                                    else :
                                        found = False
                                        for itemcur2 in itemcur[1] :
                                            #print('itemcur2',itemcur2)
                                            #print('itemcur2[0]',itemcur2[0])
                                            #print('itemcur2[1]',itemcur2[1])
                                            if itemcur[1][0] == item[67] :
                                                found = True

                                            if item[68] != '':                                                
                                                itemcur[1][1] = float(itemcur[1][1])+float(item[68])

                                        if found == False : 
                                            itemcur[1].append([item[67],item[68]])
                                else :
                                    bankdata.append([item[59],[item[67],item[68]]])
                        pp(bankdata)

                        item44=item[44].split(';')
                        len1=len(item44)
                        rowlen=len1/5
                        t1 = item44[0] 
                        t2 = item44[2] 
                        t3 = item44[3] 
                        newitem.append(t1) #12
                        newitem.append(t2) # 13
                        newitem.append(t3) # 14
                        newitem.append(item[40]) # 15
                        if len(selecteddata) :
                            newitem.append(selecteddata[0][21]) # 16
                        else:
                            newitem.append('')
                        
                        if len(selecteddata) :
                            newitem.append(selecteddata[0][26]) # 17
                        else :
                            newitem.append('')
                        
                        if len(selecteddata) :
                            newitem.append(selecteddata[0][28]) # 18
                        else:
                            newitem.append('')

                        newitem.append(item[60]) # 19
                        if len(selecteddata) :
                            newitem.append(selecteddata[0][31]) # 20 remakr shipping  request
                        else :
                            newitem.append('')

                        newitem.append('') # 21 
                        newitem.append(item[2]) # 22
                        newitem.append(item[3]) # 23

                        
                        total_db_final.append(newitem)
            
            db_data_final =[]
            db_data_final.append('NEGO예정일')  #0
            db_data_final.append('담당')  #1          
            db_data_final.append('OrderNo')  #2
            db_data_final.append('Buyer') #3
            db_data_final.append('Item') #4
            db_data_final.append('Qty')   #5               
            db_data_final.append('Nego단가')   #6
            db_data_final.append('단위') #7
            db_data_final.append('Amount') #8
            db_data_final.append('Term') #9
            db_data_final.append('Nego은행') #10
            db_data_final.append('상호') #11
            db_data_final.append('단가') #12
            db_data_final.append('Local Amount') #13
            db_data_final.append('LocalLcNo') #14
            db_data_final.append('선박회사') #15
            db_data_final.append('출항예정일') #16
            db_data_final.append('선임') #17
            db_data_final.append('서류일자') #18
            db_data_final.append('Remark') #19
            db_data_final.append('Section') #20
            db_data_final.append('Orders') #21
            db_data_final.append('OrderGubun') #22
            
            

            #print( 'total_db_final',len(total_db_final[0]) )
            #print( 'db_data_final',len(db_data_final) )
                
            return render_template('viewdb_ret_search_jugan.html', \
                                    total_db=total_db_final, db_data=db_data_final, curdata=currencydata, bankdata=bankdata)
        except TemplateNotFound:
            abort(404)


    @app.route('/save_delivery_order_db/', methods=['POST', 'GET'])
    def save_delivery_order_db():

        form_data = request.json
        ## DB
        data = DB_CONT
        ##  

        insert_values=[]
        ## save DB
        i =0 
        for item in form_data:
            if i is 0:
                item = item + form_data[1]
            insert_values.append(item)
            i += 1

        ##### Add new data to a DB

        DB_NM = 'alldata_deliveryorder'
        db_path = data[DB_NM]['path']
        table_name = data[DB_NM]['table_name']
        col_num = 30
        column_list = data[DB_NM]['column']
        col_list = ''
        for item in column_list:            
            col_list += '"' + item + '",'

        col_list = col_list[:-1]

        ## check same db
        where = 'orderno="'+str(insert_values[0])+'"'
        temp = query_db(db_path, selection_query(table_name, 'orderno' ,where))
        
        if len(temp) : 
            for j in range(col_num) :
                if (j == 0) | (j == 1):
                    continue
                
                #print("column_list[j]",column_list[j])
                #print("insert_values[j]",insert_values[j])
                #print("insert_values[0]",insert_values[0])
                #print("insert_values[1]",insert_values[1])

                update_db(db_path, update_query2(table_name, column_list[j],'orderno', 'gubun'), [insert_values[j],insert_values[0],insert_values[1]])            
        else :
            update_db(db_path, insert_query(table_name, '(' + col_list + ')', col_num), insert_values)


        # shipment main DB Save 선사 선명 포워더                     
        where_var = 'OrderNo="'+form_data[0]+'"'
        where_var2 = 'OrderGubun="' + form_data[1] + '"'
        path = 'dv/data/db/alldata.db'
        table_name = 'ALLDATA'
        select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))
        #print(select_new[0][58])
        select_new[0][58]
        containerarr = select_new[0][58].split(';')
        containerarr[5] = form_data[14]
        containerarr[6] = form_data[17]
        containerarr[8] = form_data[15]
        tablevalue = ''            
        for item_sub in containerarr:
            tablevalue += str(item_sub) + ';'
        finalitem = tablevalue
        update_db(path, update_query2(table_name, 'tableSPL_ship', 'OrderNo', 'OrderGubun'), [finalitem,form_data[0],form_data[1]])
        #select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))
        #print(select_new[0][58])        
        #print(select_new)
        #print(form_data[14]) ship company 5
        #print(form_data[15]) forward 8
        #print(form_data[17]) vessel 6
        # end shipment main DB Save 

        
        msg = 'Save complete'
        return jsonify({'msg': msg})

    @app.route('/get_delivery_order_db/', methods=['POST', 'GET'])
    def get_delivery_order_db():
        form_data = request.json
        ## DB
        data = DB_CONT
        ##  
        # alldata_ship        
        selected = form_data[0]

        path1 = data['alldata_deliveryorder']['path']
        table = 'alldata_deliveryorder'
        where = 'orderno="'+str(selected)+'"'
        selecteddata = query_db(path1,selection_query(table,'*',where))        
        # end alldata_ship
        #pp(selecteddata)
        if len(selecteddata):
            msg = 'Get data complete'
            return jsonify({'msg': msg, 'data':selecteddata})
        else :
            return jsonify({'msg': 'Fail to read data'})

    @app.route('/get_VGM_DB/', methods=['POST', 'GET'])
    def get_VGM_DB():
        data = DB_CONT
        form_data = request.json

        form_order_no = form_data
        #form_order_no = form_order_no[1:]

        form_order_no = form_order_no.replace(' ','')
        no = form_order_no.split('_')[0]
        gubun = form_order_no.split('_')[1]

        exist = 'none' # new , old
        item123 =[]
        #with open('newdb.list','r') as fh:
        #    for item in fh:
        #        item = item.strip('\n')
                
        #        if str(item).find(str(no)) != -1:
        if str(form_order_no).find(str('_')) != -1:
            item = str(form_order_no)

            exist = 'new'
            
            where_var = 'OrderNo="'+item.split('_')[0]+'"'
            where_var2 = 'OrderGubun="' + item.split('_')[1] + '"'
            path = 'dv/data/db/alldata.db'
            table_name = 'ALLDATA'
            select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))                  

            # container
            containerarr = select_new[0][58].split(';')
            container = containerarr[0]                    
            
            where_var = 'orderno="'+str(no)+str(gubun)+'"'
            booking = query_db(data['alldata_deliveryorder']['path'], selection_query(data['alldata_deliveryorder']['table_name'],'booking',where_var))


        if exist == 'new' :
            
            return jsonify({'check_new':'checked', 'no':no, 'gubun':gubun
                        ,'container':container,'book':booking})

        msg = '[error] error'
        return jsonify({'msg':msg})        

    

    @app.route('/excel_VGM/', methods=['POST','GET'])
    def excel_VGM():
        form_data=request.json

        wb = openpyxl.load_workbook(filename = 'dv/static/VGM.xlsx', read_only=False, data_only=False)

        ws = wb['SIMPLE VGM']
        ws['a4'] = form_data[0]
        ws['f4'] = form_data[1]
        wb.save('dv/static/test.xlsx')

       

        msg = 'success writing excel'
        return jsonify({'msg':msg})


    @app.route('/get_standing_payment_inst_DB/', methods=['POST', 'GET'])
    def get_standing_payment_inst_DB():
        data = DB_CONT
        form_data = request.json

        form_order_no = form_data
        #form_order_no = form_order_no[1:]

        form_order_no = form_order_no.replace(' ','')
        no = form_order_no.split('_')[0]
        gubun = form_order_no.split('_')[1]

        exist = 'none' # new , old
        
        #with open('newdb.list','r') as fh:
        #    for item in fh:
        #        item = item.strip('\n')
                
        #        if str(item).find(str(no)) != -1:
        if str(form_order_no).find(str('_')) != -1:
            item = str(form_order_no)

            exist = 'new'
            
            where_var = 'OrderNo="'+item.split('_')[0]+'"'
            where_var2 = 'OrderGubun="' + item.split('_')[1] + '"'
            path = 'dv/data/db/alldata.db'
            table_name = 'ALLDATA'
            select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))                  

            
            #where_var = 'orderno="'+str(no)+str(gubun)+'"'
            #booking = query_db(data['alldata_deliveryorder']['path'], selection_query(data['alldata_deliveryorder']['table_name'],'booking',where_var))

            # buyer #                     
            buyerselected = select_new[0][1]
            path_buyer = data['BUYER']['path']
            table_buyer = 'BUYER'
            where_buyer = '프린트Buyer="'+str(buyerselected)+'"'
            select_buyer = query_db(path_buyer,selection_query(table_buyer,'*',where_buyer))

        if exist == 'new' :
            
            return jsonify({'check_new':'checked', 'no':no, 'gubun':gubun
                        ,'select_buyer':select_buyer})

        msg = '[error] error'
        return jsonify({'msg':msg})    

    @app.route('/get_standing_payment_inst_bank/', methods=['POST', 'GET'])
    def get_standing_payment_inst_bank():
        data = DB_CONT
        form_data = request.json # keb

        # buyer #                             
        path_bank = data['inst_nego_bank']['path']
        table_bank = 'INST_NEGO_BANK'
        where_bank = 'name="'+str(form_data)+'"'
        select_bank = query_db(path_bank,selection_query(table_bank,'*',where_bank))            
        return jsonify({'select_bank':select_bank})

    @app.route('/get_standing_payment_inst_buyer/', methods=['POST', 'GET'])
    def get_standing_payment_inst_buyer():
        data = DB_CONT
        form_data = request.json # keb

        # buyer #                             
        buyerselected = form_data
        path_buyer = data['BUYER']['path']
        table_buyer = 'BUYER'
        where_buyer = '프린트Buyer="'+str(buyerselected)+'"'
        select_buyer = query_db(path_buyer,selection_query(table_buyer,'*',where_buyer))
        return jsonify({'select_buyer':select_buyer})


    
    @app.route('/get_standing_payment_inst_buyer_DB/', methods=['POST', 'GET'])
    def get_standing_payment_inst_buyer_DB():
        data = DB_CONT
        form_data = request.json

        form_order_no = form_data
        #form_order_no = form_order_no[1:]

        form_order_no = form_order_no.replace(' ','')
        no = form_order_no.split('_')[0]
        gubun = form_order_no.split('_')[1]

        exist = 'none' # new , old
        
        #with open('newdb.list','r') as fh:
        #    for item in fh:
        #        item = item.strip('\n')
                
        #        if str(item).find(str(no)) != -1:
        if str(form_order_no).find(str('_')) != -1:
            item = str(form_order_no)

            exist = 'new'
            
            where_var = 'OrderNo="'+item.split('_')[0]+'"'
            where_var2 = 'OrderGubun="' + item.split('_')[1] + '"'
            path = 'dv/data/db/alldata.db'
            table_name = 'ALLDATA'
            select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))                  

            # buyer #                     
            buyerselected = select_new[0][1]
            path_buyer = data['BUYER']['path']
            table_buyer = 'BUYER' 
            where_buyer = '프린트Buyer="'+str(buyerselected)+'"'
            select_buyer = query_db(path_buyer,selection_query(table_buyer,'*',where_buyer))

            #amount
            dataAmount = select_new[0][39]

        if exist == 'new' :
            
            return jsonify({'check_new':'checked', 'no':no, 'gubun':gubun
                        ,'select_buyer':select_buyer, 'dataAmount':dataAmount})

        msg = '[error] error'
        return jsonify({'msg':msg})    

    @app.route('/get_standing_payment_inst_buyer_bank/', methods=['POST', 'GET'])
    def get_standing_payment_inst_buyer_bank():
        data = DB_CONT
        form_data = request.json # keb

        # buyer #                             
        path_bank = data['inst_nego_bank']['path']
        table_bank = 'INST_NEGO_BANK'
        where_bank = 'name="'+str(form_data)+'"'
        select_bank = query_db(path_bank,selection_query(table_bank,'*',where_bank))            
        return jsonify({'select_bank':select_bank})

    @app.route('/get_standing_payment_inst_buyer_buyer/', methods=['POST', 'GET'])
    def get_standing_payment_inst_buyer_buyer():
        data = DB_CONT
        form_data = request.json # keb

        # buyer #                             
        buyerselected = form_data
        path_buyer = data['BUYER']['path']
        table_buyer = 'BUYER'
        where_buyer = '프린트Buyer="'+str(buyerselected)+'"'
        select_buyer = query_db(path_buyer,selection_query(table_buyer,'*',where_buyer))
        return jsonify({'select_buyer':select_buyer})

    @app.route('/set_check_company/', methods=['POST', 'GET'])
    def set_check_company():
        data = request.json 
        CHECK_COMPANY = data
        #print(CHECK_COMPANY)
        session['company'] = CHECK_COMPANY
        return jsonify({'msg':CHECK_COMPANY})
    
    @app.route('/get_check_company/', methods=['POST', 'GET'])
    def get_check_company():
        
        CHECK_COMPANY11 = session['company'] 
        return jsonify({'out':CHECK_COMPANY11})
    

    @app.route('/remakeSelect/', methods=['POST', 'GET'])
    def remakeSelect():
        data = request.json 
        
        #nego_data = get_nego_db(data)
        nego_data = []
        all_data = get_alldata_db(data)
        return jsonify({'nego_data':nego_data, 'all_data':all_data})


    @app.route('/get_id_unit/', methods=['POST', 'GET'])
    def get_id_unit():
        data = DB_CONT
        form_data = request.json

        print(form_data)

        form_order_no = form_data

        form_order_no = form_order_no.replace(' ','')
        no = form_order_no.split('_')[0]
        gubun = form_order_no.split('_')[1]

        exist = 'none' # new , old
        item123 =[]

        print(form_order_no)
        
        if str(form_order_no).find(str('_')) != -1:
            item = str(form_order_no)

            exist = 'new'
            
            where_var = 'OrderNo="'+item.split('_')[0]+'"'
            where_var2 = 'OrderGubun="' + item.split('_')[1] + '"'
            path = 'dv/data/db/alldata.db'
            table_name = 'ALLDATA'

            select_new2_list =query_db(data['alldata']['path'], selection_query(data['alldata']['table_name'], '*', where_var))

            for select_new in select_new2_list : 
                #select_new = query_db(path, selection_query_two(table_name,'*',where_var,where_var2))                  

              
                # PRC_order 36
                PRCorder = select_new[36]

              
        if exist == 'new' :
            
            return jsonify({
                        'PRCorder':PRCorder})

        msg = '[error] error'
        return jsonify({'msg':msg})
    return app
    


